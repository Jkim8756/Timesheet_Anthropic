"""
Timesheet Batch Extractor
Processes timesheet images via Claude Vision API → Excel + optional DB export
"""

import os
import json
import base64
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional
import anthropic
import openpyxl
from dotenv import load_dotenv
load_dotenv()
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("extraction.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
MODEL = "claude-sonnet-4-20250514"
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds between retries

EXTRACTION_PROMPT = """
You are extracting data from a staff timesheet image.

Return ONLY a valid JSON object — no markdown, no explanation, no preamble.

Extract the following top-level fields:
- project (string)
- business_unit (string)
- date (string, as written)
- day_of_week (string)
- weather (string or null)

Extract "frontline_staff" as an array of objects, each with:
- job_task, title, employee_name, ein,
- scheduled_start, scheduled_end, scheduled_hours,
- actual_start, lunch_out, lunch_in, actual_end, actual_hours,
- signature (string or null),
- absent (true/false/null),
- schedule_changed (true/false/null),
- confidence (0.0–1.0 — lower if handwriting was hard to read)

Extract "management_staff" as an array with the same fields.

Extract "summary":
- attendees (integer or null)
- absent_count (integer or null)

If a field is illegible or missing, use null. Never guess — use null if uncertain.
"""

# ── Helpers ───────────────────────────────────────────────────────────────────

def encode_image(path: Path) -> tuple[str, str]:
    """Return (base64_data, media_type) for an image file."""
    ext = path.suffix.lower()
    media_map = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif"
    }
    media_type = media_map.get(ext, "image/jpeg")
    with open(path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8"), media_type


def extract_sheet(client: anthropic.Anthropic, image_path: Path) -> dict:
    """Send one image to Claude and return parsed JSON."""
    b64, media_type = encode_image(image_path)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=2000,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                        {"type": "text", "text": EXTRACTION_PROMPT}
                    ]
                }]
            )
            raw = response.content[0].text.strip()
            # Strip markdown fences if present
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
            result = json.loads(raw)
            result["_source_file"] = image_path.name
            result["_extracted_at"] = datetime.now().isoformat()
            result["_status"] = "success"
            return result

        except json.JSONDecodeError as e:
            log.warning(f"[{image_path.name}] JSON parse error attempt {attempt}: {e}")
        except anthropic.RateLimitError:
            log.warning(f"[{image_path.name}] Rate limited — waiting {RETRY_DELAY * attempt}s")
            time.sleep(RETRY_DELAY * attempt)
        except Exception as e:
            log.error(f"[{image_path.name}] Error attempt {attempt}: {e}")
            time.sleep(RETRY_DELAY)

    return {
        "_source_file": image_path.name,
        "_extracted_at": datetime.now().isoformat(),
        "_status": "failed",
        "frontline_staff": [],
        "management_staff": []
    }


# ── Excel Writer ───────────────────────────────────────────────────────────────

def build_excel(all_results: list[dict], output_path: Path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Records (flat) ──────────────────────────────────────────
    ws = wb.active
    ws.title = "All Records"

    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    alt_fill = PatternFill("solid", start_color="EEF2F7", end_color="EEF2F7")
    flag_fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    fail_fill = PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0")

    columns = [
        "Source File", "Date", "Day", "Business Unit", "Project",
        "Staff Type", "Job Task", "Title", "Employee Name", "EIN",
        "Sched Start", "Sched End", "Sched Hours",
        "Actual Start", "Lunch Out", "Lunch In", "Actual End", "Actual Hours",
        "Absent", "Schedule Changed", "Confidence", "Status"
    ]

    for col_idx, col_name in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    row = 2
    for result in all_results:
        if result["_status"] == "failed":
            c = ws.cell(row=row, column=1, value=result["_source_file"])
            c.fill = fail_fill
            ws.cell(row=row, column=22, value="FAILED")
            row += 1
            continue

        meta = [
            result.get("_source_file", ""),
            result.get("date", ""),
            result.get("day_of_week", ""),
            result.get("business_unit", ""),
            result.get("project", ""),
        ]

        all_staff = (
            [("Frontline", s) for s in (result.get("frontline_staff") or [])] +
            [("Management", s) for s in (result.get("management_staff") or [])]
        )

        for staff_type, s in all_staff:
            confidence = s.get("confidence", 1.0) or 1.0
            row_fill = flag_fill if confidence < 0.8 else (alt_fill if row % 2 == 0 else None)

            values = meta + [
                staff_type,
                s.get("job_task", ""),
                s.get("title", ""),
                s.get("employee_name", ""),
                s.get("ein", ""),
                s.get("scheduled_start", ""),
                s.get("scheduled_end", ""),
                s.get("scheduled_hours", ""),
                s.get("actual_start", ""),
                s.get("lunch_out", ""),
                s.get("lunch_in", ""),
                s.get("actual_end", ""),
                s.get("actual_hours", ""),
                "Yes" if s.get("absent") else ("No" if s.get("absent") is False else ""),
                "Yes" if s.get("schedule_changed") else ("No" if s.get("schedule_changed") is False else ""),
                confidence,
                "⚠ Low Confidence" if confidence < 0.8 else "OK"
            ]

            for col_idx, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = bdr
                if row_fill:
                    c.fill = row_fill

            row += 1

    # Column widths
    widths = [20, 12, 10, 14, 22, 12, 10, 14, 22, 14,
              11, 11, 11, 11, 11, 11, 11, 11, 8, 14, 11, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Batch Extraction Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Calibri", color="1F3864")
    ws2["A3"] = "Total files processed:"
    ws2["B3"] = len(all_results)
    ws2["A4"] = "Successful:"
    ws2["B4"] = f'=COUNTIF(\'All Records\'!V:V,"OK")+COUNTIF(\'All Records\'!V:V,"⚠ Low Confidence")'
    ws2["A5"] = "Failed files:"
    ws2["B5"] = sum(1 for r in all_results if r["_status"] == "failed")
    ws2["A6"] = "Low confidence rows:"
    ws2["B6"] = '=COUNTIF(\'All Records\'!V:V,"⚠ Low Confidence")'
    ws2["A7"] = "Extracted at:"
    ws2["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for cell in ["A3", "A4", "A5", "A6", "A7"]:
        ws2[cell].font = Font(bold=True, name="Calibri", size=10)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 18

    # ── Sheet 3: Failed Files ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Failed & Review")
    ws3["A1"] = "File"
    ws3["B1"] = "Status"
    ws3["C1"] = "Extracted At"
    for cell in ["A1", "B1", "C1"]:
        ws3[cell].font = Font(bold=True, color="FFFFFF", name="Calibri")
        ws3[cell].fill = hdr_fill

    r = 2
    for result in all_results:
        if result["_status"] == "failed":
            ws3.cell(row=r, column=1, value=result["_source_file"])
            ws3.cell(row=r, column=2, value="FAILED")
            ws3.cell(row=r, column=3, value=result["_extracted_at"])
            r += 1
        else:
            for staff in (result.get("frontline_staff") or []) + (result.get("management_staff") or []):
                if (staff.get("confidence") or 1.0) < 0.8:
                    ws3.cell(row=r, column=1, value=result["_source_file"])
                    ws3.cell(row=r, column=2, value=f'Low confidence: {staff.get("employee_name","")} ({staff.get("confidence","")})')
                    ws3.cell(row=r, column=3, value=result["_extracted_at"])
                    r += 1

    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 20

    wb.save(output_path)
    log.info(f"Excel saved → {output_path}")


# ── Database Export (optional) ─────────────────────────────────────────────────

def export_to_db(all_results: list[dict], db_url: str):
    """
    Export to PostgreSQL. Requires: pip install psycopg2-binary sqlalchemy
    db_url example: postgresql://user:password@localhost:5432/timesheets
    """
    try:
        from sqlalchemy import create_engine, text
        import pandas as pd
    except ImportError:
        log.error("DB export requires: pip install psycopg2-binary sqlalchemy pandas")
        return

    rows = []
    for result in all_results:
        if result["_status"] == "failed":
            continue
        for staff_type, staff_list in [("frontline", result.get("frontline_staff") or []),
                                        ("management", result.get("management_staff") or [])]:
            for s in staff_list:
                rows.append({
                    "source_file": result.get("_source_file"),
                    "extracted_at": result.get("_extracted_at"),
                    "project": result.get("project"),
                    "business_unit": result.get("business_unit"),
                    "work_date": result.get("date"),
                    "day_of_week": result.get("day_of_week"),
                    "staff_type": staff_type,
                    "job_task": s.get("job_task"),
                    "title": s.get("title"),
                    "employee_name": s.get("employee_name"),
                    "ein": s.get("ein"),
                    "scheduled_start": s.get("scheduled_start"),
                    "scheduled_end": s.get("scheduled_end"),
                    "scheduled_hours": s.get("scheduled_hours"),
                    "actual_start": s.get("actual_start"),
                    "lunch_out": s.get("lunch_out"),
                    "lunch_in": s.get("lunch_in"),
                    "actual_end": s.get("actual_end"),
                    "actual_hours": s.get("actual_hours"),
                    "absent": s.get("absent"),
                    "schedule_changed": s.get("schedule_changed"),
                    "confidence": s.get("confidence"),
                })

    df = pd.DataFrame(rows)
    engine = create_engine(db_url)

    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS timesheet_staging (
                id SERIAL PRIMARY KEY,
                source_file VARCHAR(255),
                extracted_at TIMESTAMP,
                project VARCHAR(255),
                business_unit VARCHAR(100),
                work_date VARCHAR(50),
                day_of_week VARCHAR(20),
                staff_type VARCHAR(20),
                job_task VARCHAR(10),
                title VARCHAR(100),
                employee_name VARCHAR(150),
                ein VARCHAR(30),
                scheduled_start VARCHAR(20),
                scheduled_end VARCHAR(20),
                scheduled_hours VARCHAR(10),
                actual_start VARCHAR(20),
                lunch_out VARCHAR(20),
                lunch_in VARCHAR(20),
                actual_end VARCHAR(20),
                actual_hours VARCHAR(10),
                absent BOOLEAN,
                schedule_changed BOOLEAN,
                confidence NUMERIC(3,2),
                reviewed BOOLEAN DEFAULT FALSE
            )
        """))
        conn.commit()

    df.to_sql("timesheet_staging", engine, if_exists="append", index=False)
    log.info(f"Exported {len(df)} rows to database staging table.")


# ── Main Runner ────────────────────────────────────────────────────────────────

def run_batch(
    input_folder: str,
    output_excel: str = "timesheet_output.xlsx",
    results_json: str = "results.json",
    db_url: Optional[str] = None,
    delay_between_requests: float = 0.5
):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError("Set ANTHROPIC_API_KEY environment variable before running.")

    client = anthropic.Anthropic(api_key=api_key)
    input_path = Path(input_folder)

    images = sorted([
        f for f in input_path.iterdir()
        if f.suffix.lower() in SUPPORTED_EXTENSIONS
    ])

    if not images:
        log.warning(f"No supported images found in: {input_folder}")
        return

    log.info(f"Found {len(images)} images to process.")
    all_results = []

    for i, img_path in enumerate(images, 1):
        log.info(f"[{i}/{len(images)}] Processing: {img_path.name}")
        result = extract_sheet(client, img_path)
        all_results.append(result)

        status = result["_status"]
        staff_count = len(result.get("frontline_staff") or []) + len(result.get("management_staff") or [])
        log.info(f"  → {status.upper()} | {staff_count} staff records extracted")

        # Save incremental JSON after each file (crash recovery)
        with open(results_json, "w") as f:
            json.dump(all_results, f, indent=2)

        time.sleep(delay_between_requests)

    # Build Excel
    build_excel(all_results, Path(output_excel))

    # Optional DB export
    if db_url:
        export_to_db(all_results, db_url)

    # Final summary
    success = sum(1 for r in all_results if r["_status"] == "success")
    failed = len(all_results) - success
    log.info(f"\n{'─'*50}")
    log.info(f"Batch complete: {success} succeeded, {failed} failed")
    log.info(f"Excel output:   {output_excel}")
    log.info(f"JSON backup:    {results_json}")
    log.info(f"{'─'*50}")


# ── Entry Point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Batch timesheet extractor using Claude Vision")
    parser.add_argument("input_folder", help="Folder containing timesheet images")
    parser.add_argument("--output", default="timesheet_output.xlsx", help="Output Excel filename")
    parser.add_argument("--json", default="results.json", help="JSON backup filename")
    parser.add_argument("--db", default=None, help="PostgreSQL URL for DB export (optional)")
    parser.add_argument("--delay", type=float, default=0.5, help="Seconds between API calls (default: 0.5)")
    args = parser.parse_args()

    run_batch(
        input_folder=args.input_folder,
        output_excel=args.output,
        results_json=args.json,
        db_url=args.db,
        delay_between_requests=args.delay
    )