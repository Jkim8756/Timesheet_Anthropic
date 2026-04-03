#!/usr/bin/env python3
"""
Step 1 – Extract timesheet data from PDF files using Claude Vision.

Usage:
    python src/extract.py
    python src/extract.py --model claude-opus-4-20250514
    python src/extract.py --engine google
    python src/extract.py --output my_output.xlsx
    python src/extract.py --from-json output/json/extraction_<ts>.json

Outputs:
    Supabase PostgreSQL  timesheet_entries  (primary store)
    output/excel/<source>_extracted.xlsx    (per-source Excel)
    output/excel/timesheet_output.xlsx      (master Excel)
    output/json/extraction_<ts>.json        (backup)
    output/csv/extraction_<ts>.csv          (backup)
    logs/extract.log

Requirements:  pip install -r requirements.txt
               Set ANTHROPIC_API_KEY and DATABASE_URL in .env
               For --engine google: set GOOGLE_APPLICATION_CREDENTIALS in .env
"""

import base64
import io
import json
import logging
import os
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import anthropic
import openpyxl
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter

# Optional dependencies — only needed for --engine google
try:
    import fitz as _fitz          # PyMuPDF: render PDF pages to PNG
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False

try:
    from google.cloud import vision as _vision
    _GOOGLE_AVAILABLE = True
except ImportError:
    _GOOGLE_AVAILABLE = False

load_dotenv()

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).resolve().parent.parent
INPUT_DIR  = BASE_DIR / "input"
DONE_DIR   = INPUT_DIR / "processed"
EXCEL_DIR  = BASE_DIR / "output" / "excel"
JSON_DIR   = BASE_DIR / "output" / "json"
CSV_DIR    = BASE_DIR / "output" / "csv"
LOG_DIR    = BASE_DIR / "logs"

for _d in (INPUT_DIR, DONE_DIR, EXCEL_DIR, JSON_DIR, CSV_DIR, LOG_DIR):
    _d.mkdir(parents=True, exist_ok=True)

# ── Model / pricing ────────────────────────────────────────────────────────────
# Change DEFAULT_MODEL to switch models for future runs.
DEFAULT_MODEL = "claude-sonnet-4-20250514"

# Pricing per million tokens (USD). Update if Anthropic changes rates.
# https://www.anthropic.com/pricing
MODEL_PRICING = {
    "claude-opus-4":    (15.00, 75.00),
    "claude-sonnet-4":  ( 3.00, 15.00),
    "claude-haiku-4":   ( 0.80,  4.00),
    "claude-haiku-3":   ( 0.25,  1.25),
    "default":          ( 3.00, 15.00),
}

# Token estimate per page for pre-flight cost estimate only
_EST_INPUT_PER_PAGE  = 1600
_EST_OUTPUT_PER_PAGE = 600

MAX_RETRIES = 3
RETRY_DELAY = 5   # seconds between retry attempts

# Google Cloud Vision cost (DOCUMENT_TEXT_DETECTION)
GOOGLE_COST_PER_PAGE = 0.0015  # USD per page

# Token estimates for Google engine (Claude receives text, not a PDF document — smaller)
_EST_INPUT_PER_PAGE_GOOGLE  = 800
_EST_OUTPUT_PER_PAGE_GOOGLE = 600

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "extract.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ── Claude client ──────────────────────────────────────────────────────────────
client = anthropic.Anthropic()   # reads ANTHROPIC_API_KEY from env / .env

# ── Extraction prompt ──────────────────────────────────────────────────────────
EXTRACTION_PROMPT = """
You are extracting data from a staff timesheet image (one page only).

Return ONLY a valid JSON object - no markdown, no explanation, no preamble.

Extract the following top-level fields:
- project (string)
- business_unit (string)
- date (string, as written)
- day_of_week (string)
- weather (string or null)

Extract "frontline_staff" as an array of objects, each with:
- job_task, title, employee_name, ein,
- scheduled_start, scheduled_end, scheduled_hours,
- actual_start, lunch_out, lunch_in, actual_end, actual_hours,
- signature (string or null),
- absent (true/false/null),
- schedule_changed (true/false/null),
- confidence (0.0-1.0 - lower if handwriting was hard to read)

Extract "management_staff" as an array with the same fields.

Extract "summary":
- attendees (integer or null)
- absent_count (integer or null)

If a field is illegible or missing, use null. Never guess - use null if uncertain.
"""

# Used by --engine google: Claude receives raw OCR text from Google Vision and
# structures it into the same JSON schema as EXTRACTION_PROMPT.
GOOGLE_STRUCTURE_PROMPT = """\
You are structuring raw OCR text extracted from a staff timesheet (one page only).
The text below was produced by Google Cloud Vision and may contain minor OCR artifacts.

Return ONLY a valid JSON object - no markdown, no explanation, no preamble.

Extract the following top-level fields:
- project (string)
- business_unit (string)
- date (string, as written)
- day_of_week (string)
- weather (string or null)

Extract "frontline_staff" as an array of objects, each with:
- job_task, title, employee_name, ein,
- scheduled_start, scheduled_end, scheduled_hours,
- actual_start, lunch_out, lunch_in, actual_end, actual_hours,
- signature (string or null),
- absent (true/false/null),
- schedule_changed (true/false/null),
- confidence (0.0-1.0 - lower if the OCR text was ambiguous or incomplete)

Extract "management_staff" as an array with the same fields.

Extract "summary":
- attendees (integer or null)
- absent_count (integer or null)

If a field is missing or ambiguous, use null. Never guess.

OCR TEXT:
{ocr_text}
"""

# ── PostgreSQL schema ──────────────────────────────────────────────────────────
_DDL = """
CREATE TABLE IF NOT EXISTS timesheet_entries (
    id                 SERIAL        PRIMARY KEY,
    source_file        TEXT,
    page_number        INTEGER,
    extracted_at       TIMESTAMPTZ,
    extraction_status  TEXT,
    model_used         TEXT,
    project            TEXT,
    business_unit      TEXT,
    work_date          TEXT,
    day_of_week        TEXT,
    weather            TEXT,
    staff_type         TEXT,
    job_task           TEXT,
    title              TEXT,
    employee_name      TEXT,
    ein                TEXT,
    scheduled_start    TEXT,
    scheduled_end      TEXT,
    scheduled_hours    TEXT,
    actual_start       TEXT,
    lunch_out          TEXT,
    lunch_in           TEXT,
    actual_end         TEXT,
    actual_hours       TEXT,
    absent             BOOLEAN,
    schedule_changed   BOOLEAN,
    confidence         NUMERIC(3,2),
    signature          TEXT,
    raw_json           JSONB
);
"""

_SQL_COLS = [
    "source_file", "page_number", "extracted_at", "extraction_status", "model_used",
    "project", "business_unit", "work_date", "day_of_week", "weather",
    "staff_type", "job_task", "title", "employee_name", "ein",
    "scheduled_start", "scheduled_end", "scheduled_hours",
    "actual_start", "lunch_out", "lunch_in", "actual_end", "actual_hours",
    "absent", "schedule_changed", "confidence", "signature", "raw_json",
]

# ── Excel column definitions ───────────────────────────────────────────────────
STAFF_COLUMNS = [
    "Source File", "Page", "Date", "Day", "Business Unit", "Project",
    "Staff Type", "Job Task", "Title", "Employee Name", "EIN",
    "Sched Start", "Sched End", "Sched Hours",
    "Actual Start", "Lunch Out", "Lunch In", "Actual End", "Actual Hours",
    "Absent", "Schedule Changed", "Confidence", "Status",
]

STAFF_COL_WIDTHS = [
    22, 6, 12, 10, 14, 22, 12, 10, 14,
    22, 14, 11, 11, 11, 11, 11, 11, 11, 11,
    8, 16, 11, 16,
]


# ── Cost helpers ───────────────────────────────────────────────────────────────

def _get_pricing(model: str) -> tuple[float, float]:
    for key, pricing in MODEL_PRICING.items():
        if key in model:
            return pricing
    return MODEL_PRICING["default"]


def _estimate_cost(total_pages: int, model: str, engine: str = "claude") -> float:
    inp, out = _get_pricing(model)
    if engine == "google":
        google_cost = total_pages * GOOGLE_COST_PER_PAGE
        claude_cost = (total_pages * _EST_INPUT_PER_PAGE_GOOGLE / 1_000_000 * inp +
                       total_pages * _EST_OUTPUT_PER_PAGE_GOOGLE / 1_000_000 * out)
        return google_cost + claude_cost
    return (total_pages * _EST_INPUT_PER_PAGE / 1_000_000 * inp +
            total_pages * _EST_OUTPUT_PER_PAGE / 1_000_000 * out)


def _actual_cost(usage: dict, model: str) -> float:
    inp, out = _get_pricing(model)
    return (usage.get("input_tokens", 0) / 1_000_000 * inp +
            usage.get("output_tokens", 0) / 1_000_000 * out)


# ── PDF helpers ────────────────────────────────────────────────────────────────

def _pdf_to_pages(pdf_path: Path) -> list[dict]:
    """Split PDF into individual pages; return list of {page_number, b64, media_type}."""
    reader = PdfReader(str(pdf_path))
    pages = []
    for i, page in enumerate(reader.pages, 1):
        writer = PdfWriter()
        writer.add_page(page)
        buf = io.BytesIO()
        writer.write(buf)
        b64 = base64.standard_b64encode(buf.getvalue()).decode()
        pages.append({
            "page_number": i,
            "image_b64":   b64,
            "media_type":  "application/pdf",
        })
    return pages


# ── Interactive file selection ─────────────────────────────────────────────────

def _select_files(model: str, engine: Optional[str] = None) -> tuple[list[Path], str]:
    """
    Interactive file + engine selection.
    Returns (selected_files, engine_name).
    If engine is passed (e.g. from --engine flag) the engine prompt is skipped.
    """
    candidates = sorted(INPUT_DIR.glob("*.pdf"))
    if not candidates:
        log.warning("No PDF files found in %s", INPUT_DIR)
        sys.exit(0)

    # Page count per file
    page_counts = {}
    for f in candidates:
        try:
            page_counts[f] = len(PdfReader(str(f)).pages)
        except Exception:
            page_counts[f] = "?"

    print("\n" + "=" * 62)
    print("  TIMESHEET EXTRACTOR – FILE SELECTION")
    print("=" * 62)
    print(f"  Model  : {model}")
    print()
    for i, f in enumerate(candidates, 1):
        kb = f.stat().st_size / 1024
        print(f"  [{i}] {f.name:<42} {page_counts[f]:>3} pages  ({kb:.0f} KB)")

    print("\nSelect files (e.g. 1,2  or  all  or  q to quit):")
    while True:
        raw = input("  > ").strip().lower()
        if raw == "q":
            print("Exiting.")
            sys.exit(0)
        if raw == "all":
            selected = candidates
            break
        try:
            indices  = [int(x.strip()) - 1 for x in raw.split(",")]
            selected = [candidates[i] for i in indices if 0 <= i < len(candidates)]
            if selected:
                break
            print("  No valid selections. Try again.")
        except ValueError:
            print("  Invalid input.")

    total_pages = sum(page_counts[f] for f in selected if isinstance(page_counts[f], int))
    est_claude = _estimate_cost(total_pages, model, "claude")
    est_google = _estimate_cost(total_pages, model, "google")

    print("\n" + "-" * 62)
    print("  FILES SELECTED:")
    for f in selected:
        print(f"    • {f.name}  ({page_counts[f]} pages)")
    print(f"\n  Total pages : {total_pages}")
    print()
    print("  Estimated cost:")
    print(f"    [1] Claude  (PDF → Claude direct)              ${est_claude:.4f} USD")
    print(f"    [2] Google  (Google Vision OCR → Claude)       ${est_google:.4f} USD")
    print("-" * 62)

    # Engine selection — skip prompt if already provided via CLI flag
    if engine is None:
        print("\nSelect engine (1 = Claude, 2 = Google, q to quit):")
        while True:
            raw = input("  > ").strip().lower()
            if raw == "q":
                print("Cancelled.")
                sys.exit(0)
            if raw in ("1", "claude"):
                engine = "claude"
                break
            if raw in ("2", "google"):
                engine = "google"
                break
            print("  Enter 1 or 2.")
    else:
        print(f"\n  Engine : {engine}  (set via --engine flag)")

    print(f"\n  Engine selected : {engine}")
    print(f"  Estimated cost  : ${est_claude if engine == 'claude' else est_google:.4f} USD")
    print("\nProceed? (y/n):")
    if input("  > ").strip().lower() != "y":
        print("Cancelled.")
        sys.exit(0)

    return selected, engine


# ── Claude API call ────────────────────────────────────────────────────────────

def _call_claude(page: dict, label: str, model: str) -> tuple[Optional[dict], dict]:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.messages.create(
                model=model,
                max_tokens=8000,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": page["media_type"],
                                "data": page["image_b64"],
                            },
                        },
                        {"type": "text", "text": EXTRACTION_PROMPT},
                    ],
                }],
            )

            raw = response.content[0].text.strip()

            # Strip markdown fences if Claude adds them despite the prompt
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                raw = raw.strip()

            data = json.loads(raw)
            usage = {
                "input_tokens":  response.usage.input_tokens,
                "output_tokens": response.usage.output_tokens,
            }
            return data, usage

        except json.JSONDecodeError:
            log.warning("[%s] JSON parse failed (attempt %d)", label, attempt)
            if attempt == MAX_RETRIES:
                return None, {}

        except anthropic.RateLimitError:
            wait = RETRY_DELAY * attempt
            log.warning("[%s] Rate limited — waiting %ds", label, wait)
            time.sleep(wait)

        except Exception as exc:
            log.warning("[%s] Error attempt %d: %s", label, attempt, exc)
            time.sleep(RETRY_DELAY)

    return None, {}


# ── Google Vision + Claude structuring call ────────────────────────────────────

def _call_google(page: dict, label: str, model: str) -> tuple[Optional[dict], dict]:
    """Render PDF page to PNG → Google Vision OCR → Claude JSON structuring."""
    if not _FITZ_AVAILABLE:
        raise RuntimeError(
            "PyMuPDF is required for --engine google.\n"
            "  pip install PyMuPDF"
        )
    if not _GOOGLE_AVAILABLE:
        raise RuntimeError(
            "google-cloud-vision is required for --engine google.\n"
            "  pip install google-cloud-vision"
        )

    gv_client         = _vision.ImageAnnotatorClient()
    google_pages_used = 0

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            # 1. Decode the single-page PDF
            pdf_bytes = base64.standard_b64decode(page["image_b64"])

            # 2. Render to PNG via PyMuPDF (2× zoom ≈ 144 DPI)
            pdf_doc   = _fitz.open(stream=pdf_bytes, filetype="pdf")
            fitz_page = pdf_doc[0]
            pix       = fitz_page.get_pixmap(matrix=_fitz.Matrix(2.0, 2.0))
            png_bytes = pix.tobytes("png")
            pdf_doc.close()

            # 3. Google Vision DOCUMENT_TEXT_DETECTION
            gv_image          = _vision.Image(content=png_bytes)
            gv_resp           = gv_client.document_text_detection(image=gv_image)
            google_pages_used += 1
            ocr_text          = gv_resp.full_text_annotation.text

            if not ocr_text.strip():
                log.warning("[%s] Google Vision returned empty text (attempt %d)", label, attempt)
                time.sleep(RETRY_DELAY)
                continue

            log.debug("[%s] Google Vision OCR: %d chars", label, len(ocr_text))

            # 4. Claude structures the raw OCR text into the JSON schema
            claude_resp = client.messages.create(
                model=model,
                max_tokens=8000,
                messages=[{
                    "role": "user",
                    "content": GOOGLE_STRUCTURE_PROMPT.format(ocr_text=ocr_text),
                }],
            )

            raw = claude_resp.content[0].text.strip()
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                raw = raw.strip()

            data  = json.loads(raw)
            usage = {
                "input_tokens":  claude_resp.usage.input_tokens,
                "output_tokens": claude_resp.usage.output_tokens,
                "google_pages":  google_pages_used,
            }
            return data, usage

        except json.JSONDecodeError:
            log.warning("[%s] JSON parse failed (attempt %d)", label, attempt)
            if attempt == MAX_RETRIES:
                return None, {"input_tokens": 0, "output_tokens": 0, "google_pages": google_pages_used}

        except Exception as exc:
            log.warning("[%s] Error attempt %d: %s", label, attempt, exc)
            time.sleep(RETRY_DELAY)

    return None, {"input_tokens": 0, "output_tokens": 0, "google_pages": google_pages_used}


# ── DB helpers ─────────────────────────────────────────────────────────────────

def _get_db_conn() -> psycopg2.extensions.connection:
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise RuntimeError("DATABASE_URL is not set. Add it to your .env file.")
    return psycopg2.connect(url)


def _ensure_table(conn: psycopg2.extensions.connection) -> None:
    # Migration: add any columns that didn't exist in earlier schema versions.
    # ADD COLUMN IF NOT EXISTS is safe to run repeatedly.
    _MIGRATIONS = """
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS extraction_status TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS model_used         TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS weather             TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS staff_type          TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS job_task            TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS employee_name       TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS ein                 TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS scheduled_start     TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS scheduled_end       TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS scheduled_hours     TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS actual_start        TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS actual_end          TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS actual_hours        TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS schedule_changed    BOOLEAN;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS confidence          NUMERIC(3,2);
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS signature           TEXT;
        ALTER TABLE timesheet_entries ADD COLUMN IF NOT EXISTS work_date           TEXT;
    """
    with conn.cursor() as cur:
        cur.execute(_DDL)
        for stmt in _MIGRATIONS.strip().split(";"):
            stmt = stmt.strip()
            if stmt:
                cur.execute(stmt)
    conn.commit()
    log.info("DB schema verified/migrated.")


def _save_to_db(conn: psycopg2.extensions.connection, results: list[dict], model: str) -> None:
    """Flatten page-level results into staff rows and insert into Supabase."""
    if not results:
        return

    placeholders = ", ".join(["%s"] * len(_SQL_COLS))
    sql = (
        f"INSERT INTO timesheet_entries ({', '.join(_SQL_COLS)}) "
        f"VALUES ({placeholders})"
    )
    now = datetime.now(timezone.utc).isoformat()

    with conn.cursor() as cur:
        for result in results:
            status = result.get("_status", "failed")
            base = {
                "source_file":       result.get("_source_file", ""),
                "page_number":       result.get("_page"),
                "extracted_at":      now,
                "extraction_status": status,
                "model_used":        model,
                "project":           result.get("project"),
                "business_unit":     result.get("business_unit"),
                "work_date":         result.get("date"),
                "day_of_week":       result.get("day_of_week"),
                "weather":           result.get("weather"),
            }

            if status == "failed":
                row = {**base, **{c: None for c in _SQL_COLS if c not in base}}
                row["raw_json"] = json.dumps(result, default=str)
                cur.execute(sql, [row.get(c) for c in _SQL_COLS])
                continue

            all_staff = (
                [("Frontline",  s) for s in (result.get("frontline_staff")  or [])] +
                [("Management", s) for s in (result.get("management_staff") or [])]
            )

            for staff_type, s in all_staff:
                row = {
                    **base,
                    "staff_type":       staff_type,
                    "job_task":         s.get("job_task"),
                    "title":            s.get("title"),
                    "employee_name":    s.get("employee_name"),
                    "ein":              s.get("ein"),
                    "scheduled_start":  s.get("scheduled_start"),
                    "scheduled_end":    s.get("scheduled_end"),
                    "scheduled_hours":  s.get("scheduled_hours"),
                    "actual_start":     s.get("actual_start"),
                    "lunch_out":        s.get("lunch_out"),
                    "lunch_in":         s.get("lunch_in"),
                    "actual_end":       s.get("actual_end"),
                    "actual_hours":     s.get("actual_hours"),
                    "absent":           s.get("absent"),
                    "schedule_changed": s.get("schedule_changed"),
                    "confidence":       s.get("confidence"),
                    "signature":        s.get("signature"),
                    "raw_json":         json.dumps(result, default=str),
                }
                cur.execute(sql, [row.get(c) for c in _SQL_COLS])

    conn.commit()


# ── Excel styles ───────────────────────────────────────────────────────────────

def _make_styles() -> dict:
    thin = Side(style="thin")
    return {
        "bdr":          Border(left=thin, right=thin, top=thin, bottom=thin),
        "hdr_fill":     PatternFill("solid", start_color="1F3864"),
        "alt_fill":     PatternFill("solid", start_color="EEF2F7"),
        "flag_fill":    PatternFill("solid", start_color="FFF2CC"),
        "fail_fill":    PatternFill("solid", start_color="FFE0E0"),
        "ok_fill":      PatternFill("solid", start_color="E2EFDA"),
    }


def _style_header_row(ws, styles: dict, row: int = 1) -> None:
    for col_idx, col_name in enumerate(STAFF_COLUMNS, 1):
        c = ws.cell(row=row, column=col_idx, value=col_name)
        c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill      = styles["hdr_fill"]
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = styles["bdr"]
    ws.row_dimensions[row].height = 28


def _staff_row_values(result: dict, staff_type: str, s: dict) -> list:
    conf = s.get("confidence", 1.0) or 1.0
    return [
        result.get("_source_file", ""),
        result.get("_page", ""),
        result.get("date", ""),
        result.get("day_of_week", ""),
        result.get("business_unit", ""),
        result.get("project", ""),
        staff_type,
        s.get("job_task", ""),
        s.get("title", ""),
        s.get("employee_name", ""),
        s.get("ein", ""),
        s.get("scheduled_start", ""),
        s.get("scheduled_end", ""),
        s.get("scheduled_hours", ""),
        s.get("actual_start", ""),
        s.get("lunch_out", ""),
        s.get("lunch_in", ""),
        s.get("actual_end", ""),
        s.get("actual_hours", ""),
        "Yes" if s.get("absent") is True else ("No" if s.get("absent") is False else ""),
        "Yes" if s.get("schedule_changed") is True else ("No" if s.get("schedule_changed") is False else ""),
        conf,
        "Low Confidence" if conf < 0.8 else "OK",
    ]


def _write_staff_rows(ws, all_results: list[dict], styles: dict, start_row: int = 2) -> int:
    row = start_row
    for result in all_results:
        if result["_status"] == "failed":
            c = ws.cell(row=row, column=1, value=result.get("_source_file", ""))
            c.fill = styles["fail_fill"]
            ws.cell(row=row, column=len(STAFF_COLUMNS), value="FAILED")
            row += 1
            continue

        all_staff = (
            [("Frontline",  s) for s in (result.get("frontline_staff")  or [])] +
            [("Management", s) for s in (result.get("management_staff") or [])]
        )

        for staff_type, s in all_staff:
            conf      = s.get("confidence", 1.0) or 1.0
            row_fill  = styles["flag_fill"] if conf < 0.8 else (styles["alt_fill"] if row % 2 == 0 else None)
            values    = _staff_row_values(result, staff_type, s)

            for col_idx, val in enumerate(values, 1):
                c           = ws.cell(row=row, column=col_idx, value=val)
                c.font      = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = styles["bdr"]
                if row_fill:
                    c.fill = row_fill
            row += 1
    return row


# ── Excel builders ─────────────────────────────────────────────────────────────

def _build_per_source_excel(file_results: list[dict], output_path: Path,
                             model: str, usage: dict) -> None:
    wb     = openpyxl.Workbook()
    styles = _make_styles()

    # Flat sheet
    ws_flat       = wb.active
    ws_flat.title = "All Records"
    _style_header_row(ws_flat, styles)
    ws_flat.freeze_panes = "A2"
    _write_staff_rows(ws_flat, file_results, styles)
    for i, w in enumerate(STAFF_COL_WIDTHS, 1):
        ws_flat.column_dimensions[get_column_letter(i)].width = w

    # Per-page sheets
    for result in file_results:
        page_num   = result.get("_page", 1)
        raw_name   = f"{Path(result.get('_source_file', 'file')).stem}_p{page_num}"
        sheet_name = raw_name[:31]
        for ch in r'/\*?[]:|':
            sheet_name = sheet_name.replace(ch, "")

        ws_p = wb.create_sheet(title=sheet_name)
        meta = [
            ("Project",       result.get("project", "")),
            ("Business Unit", result.get("business_unit", "")),
            ("Date",          result.get("date", "")),
            ("Day",           result.get("day_of_week", "")),
            ("Weather",       result.get("weather", "")),
            ("Page",          page_num),
            ("Status",        result.get("_status", "").upper()),
        ]
        for r_idx, (label, val) in enumerate(meta, 1):
            ws_p.cell(row=r_idx, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
            ws_p.cell(row=r_idx, column=2, value=val).font   = Font(name="Calibri", size=10)
        ws_p.column_dimensions["A"].width = 16
        ws_p.column_dimensions["B"].width = 24

        _style_header_row(ws_p, styles, row=10)
        ws_p.freeze_panes = "A11"
        last_row = _write_staff_rows(ws_p, [result], styles, start_row=11)
        for i, w in enumerate(STAFF_COL_WIDTHS, 1):
            ws_p.column_dimensions[get_column_letter(i)].width = w

        summary = result.get("summary") or {}
        ws_p.cell(row=last_row + 1, column=1, value="Attendees").font = Font(bold=True, name="Calibri")
        ws_p.cell(row=last_row + 1, column=2, value=summary.get("attendees"))
        ws_p.cell(row=last_row + 2, column=1, value="Absent").font    = Font(bold=True, name="Calibri")
        ws_p.cell(row=last_row + 2, column=2, value=summary.get("absent_count"))

    # Summary sheet
    _build_summary_sheet(wb, file_results, model, usage)

    wb.save(output_path)
    log.info("Excel saved → %s", output_path.name)


def _build_master_excel(all_results: list[dict], output_path: Path,
                         model: str, usage: dict, source_files: list[str]) -> None:
    wb     = openpyxl.Workbook()
    styles = _make_styles()

    ws_flat       = wb.active
    ws_flat.title = "All Records"
    _style_header_row(ws_flat, styles)
    ws_flat.freeze_panes = "A2"
    _write_staff_rows(ws_flat, all_results, styles)
    for i, w in enumerate(STAFF_COL_WIDTHS, 1):
        ws_flat.column_dimensions[get_column_letter(i)].width = w

    _build_summary_sheet(wb, all_results, model, usage, source_files)

    wb.save(output_path)
    log.info("Master Excel saved → %s", output_path.name)


def _build_summary_sheet(wb: openpyxl.Workbook, all_results: list[dict],
                          model: str, usage: dict,
                          source_files: Optional[list[str]] = None) -> None:
    styles   = _make_styles()
    ws_sum   = wb.create_sheet("Summary")
    ws_sum["A1"] = "Extraction Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, name="Calibri", color="1F3864")

    total_pages  = len(all_results)
    success      = sum(1 for r in all_results if r["_status"] == "success")
    failed       = total_pages - success
    total_staff  = sum(
        len(r.get("frontline_staff") or []) + len(r.get("management_staff") or [])
        for r in all_results
    )
    low_conf = sum(
        1 for r in all_results
        for s in (r.get("frontline_staff") or []) + (r.get("management_staff") or [])
        if (s.get("confidence") or 1.0) < 0.8
    )
    cost = _actual_cost(usage, model)

    rows = [
        ("", ""),
        ("Model used",            model),
        ("Total pages processed", total_pages),
        ("Successful",            success),
        ("Failed",                failed),
        ("Total staff records",   total_staff),
        ("Low confidence rows",   low_conf),
        ("Input tokens",          usage.get("input_tokens", "N/A")),
        ("Output tokens",         usage.get("output_tokens", "N/A")),
        ("Actual cost (USD)",     f"${cost:.4f}"),
        ("Extracted at",          datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Color Legend",          ""),
        ("Yellow row",            "Low confidence (<80%) — review recommended"),
        ("Red row",               "Failed — no data extracted"),
        ("Green row",             "OK"),
    ]
    if source_files:
        rows.insert(2, ("Source files", ", ".join(source_files)))

    legend_fills = {
        "Yellow row": styles["flag_fill"],
        "Red row":    styles["fail_fill"],
        "Green row":  styles["ok_fill"],
    }

    for r_idx, (label, val) in enumerate(rows, 2):
        c = ws_sum.cell(row=r_idx, column=1, value=label)
        v = ws_sum.cell(row=r_idx, column=2, value=val)
        c.font = Font(bold=True, name="Calibri", size=10)
        v.font = Font(name="Calibri", size=10)
        if label in legend_fills:
            c.fill = v.fill = legend_fills[label]

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 50
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))


# ── Per-file processor ─────────────────────────────────────────────────────────

def _process_pdf(pdf_path: Path, model: str, delay: float,
                 engine: str = "claude") -> tuple[list[dict], dict]:
    log.info("Processing: %s  [engine: %s]", pdf_path.name, engine)
    pages      = _pdf_to_pages(pdf_path)
    results    = []
    file_usage = {"input_tokens": 0, "output_tokens": 0, "google_pages": 0}

    for page in pages:
        pnum  = page["page_number"]
        label = f"{pdf_path.name} p{pnum}"
        log.info("  Page %d / %d ...", pnum, len(pages))

        if engine == "google":
            data, usage = _call_google(page, label, model)
        else:
            data, usage = _call_claude(page, label, model)

        if data is None:
            result = {
                "_source_file": pdf_path.name,
                "_page":        pnum,
                "_extracted_at": datetime.now(timezone.utc).isoformat(),
                "_status":      "failed",
                "frontline_staff":  [],
                "management_staff": [],
            }
        else:
            result = data
            result["_source_file"] = pdf_path.name
            result["_page"]        = pnum
            result["_extracted_at"] = datetime.now(timezone.utc).isoformat()
            result["_status"]      = "success"

        n_staff = (len(result.get("frontline_staff") or []) +
                   len(result.get("management_staff") or []))
        cost = _actual_cost(usage, model)

        file_usage["input_tokens"]  += usage.get("input_tokens", 0)
        file_usage["output_tokens"] += usage.get("output_tokens", 0)
        file_usage["google_pages"]  += usage.get("google_pages", 0)

        log.info(
            "  → %s  |  %d staff  |  %d in / %d out tokens  |  $%.4f",
            result["_status"].upper(), n_staff,
            usage.get("input_tokens", 0), usage.get("output_tokens", 0), cost,
        )

        results.append(result)
        time.sleep(delay)

    return results, file_usage


# ── JSON / CSV helpers ─────────────────────────────────────────────────────────

def _save_json(results: list[dict], path: Path) -> None:
    path.write_text(json.dumps(results, indent=2, default=str), encoding="utf-8")
    log.info("JSON saved → %s", path.relative_to(BASE_DIR))


def _save_csv(results: list[dict], path: Path) -> None:
    flat_rows = []
    for r in results:
        for staff_type, staff_list in [
            ("Frontline",  r.get("frontline_staff")  or []),
            ("Management", r.get("management_staff") or []),
        ]:
            for s in staff_list:
                flat_rows.append({
                    "source_file":  r.get("_source_file"),
                    "page":         r.get("_page"),
                    "date":         r.get("date"),
                    "day":          r.get("day_of_week"),
                    "business_unit": r.get("business_unit"),
                    "project":      r.get("project"),
                    "staff_type":   staff_type,
                    **s,
                })
    if flat_rows:
        pd.DataFrame(flat_rows).to_csv(path, index=False)
        log.info("CSV  saved → %s", path.relative_to(BASE_DIR))


def _write_run_log(entry: dict) -> None:
    """Append a single run summary entry to logs/runs.log as a JSON line."""
    run_log_path = LOG_DIR / "runs.log"
    with run_log_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, default=str) + "\n")


def _file_info(path: Path) -> dict:
    """Return name, relative path, and size in KB for a file."""
    if path.exists():
        return {
            "name":     path.name,
            "path":     str(path.relative_to(BASE_DIR)),
            "size_kb":  round(path.stat().st_size / 1024, 1),
        }
    return {"name": path.name, "path": str(path.relative_to(BASE_DIR)), "size_kb": None}


def load_json_for_db_import(json_path: Path, model: str) -> None:
    """Re-import a previously saved JSON file into the DB without re-calling the API."""
    log.info("Importing from JSON: %s", json_path)
    results = json.loads(json_path.read_text(encoding="utf-8"))
    conn = _get_db_conn()
    _ensure_table(conn)
    _save_to_db(conn, results, model)
    conn.close()
    log.info("DB import complete — %d page result(s) imported.", len(results))


# ── Entry point ────────────────────────────────────────────────────────────────

def main(model: str = DEFAULT_MODEL, output_excel: str = "timesheet_output.xlsx",
         delay: float = 0.5, engine: Optional[str] = None) -> None:

    selected, engine = _select_files(model, engine)
    ts             = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    all_results    = []
    all_sources    = []
    combined_usage = {"input_tokens": 0, "output_tokens": 0, "google_pages": 0}

    # ── Phase 1: API calls + local saves (JSON/CSV/Excel) ──────────────────────
    # JSON and CSV are written first. If the DB export later fails,
    # re-run with --from-json to import without paying for API calls again.

    for pdf_path in selected:
        file_results, file_usage = _process_pdf(pdf_path, model, delay, engine)

        # Checkpoint: save this file's JSON immediately after API calls finish
        stem           = pdf_path.stem.replace(" ", "_")
        file_json_path = JSON_DIR / f"{stem}_{ts}.json"
        _save_json(file_results, file_json_path)

        # Per-source Excel
        per_path = EXCEL_DIR / f"{stem}_extracted.xlsx"
        _build_per_source_excel(file_results, per_path, model, file_usage)

        all_results.extend(file_results)
        all_sources.append(pdf_path.name)
        combined_usage["input_tokens"]  += file_usage["input_tokens"]
        combined_usage["output_tokens"] += file_usage["output_tokens"]
        combined_usage["google_pages"]  += file_usage["google_pages"]

        # Archive PDF
        dest = DONE_DIR / pdf_path.name
        pdf_path.rename(dest)
        log.info("Archived: %s → input/processed/", pdf_path.name)

    # Master JSON + CSV (all files combined)
    master_json_path = JSON_DIR / f"extraction_{ts}.json"
    _save_json(all_results, master_json_path)
    _save_csv(all_results, CSV_DIR / f"extraction_{ts}.csv")

    # Master Excel
    master_path = EXCEL_DIR / output_excel
    _build_master_excel(all_results, master_path, model, combined_usage, all_sources)

    log.info("All API calls done. JSON and CSV saved. Starting DB export...")

    google_cost = combined_usage["google_pages"] * GOOGLE_COST_PER_PAGE
    cost        = _actual_cost(combined_usage, model) + google_cost
    pages_total = len(all_results)
    failed_pages = sum(1 for r in all_results if r.get("_status") == "failed")
    csv_path     = CSV_DIR / f"extraction_{ts}.csv"

    # ── Phase 2: DB export ─────────────────────────────────────────────────────
    db_status  = "success"
    db_error   = None
    try:
        conn = _get_db_conn()
        _ensure_table(conn)
        _save_to_db(conn, all_results, model)
        conn.close()
        log.info("DB export complete.")
    except Exception as exc:
        db_status = "failed"
        db_error  = str(exc)
        log.error("DB export failed: %s", exc)
        log.error("Your data is safe. Re-run DB import with:")
        log.error("  python src/extract.py --from-json \"%s\"", master_json_path)

    # ── Run log ────────────────────────────────────────────────────────────────
    run_entry = {
        "run_at":        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "status":        "success" if db_status == "success" and failed_pages == 0
                         else "partial" if db_status == "success"
                         else "db_failed",
        "engine":        engine,
        "model":         model,
        "input_files":   [
            {
                "name":  pdf.name,
                "pages": sum(1 for r in all_results if r.get("_source_file") == pdf.name),
                "failed_pages": sum(
                    1 for r in all_results
                    if r.get("_source_file") == pdf.name and r.get("_status") == "failed"
                ),
            }
            for pdf in selected
        ],
        "total_pages":   pages_total,
        "failed_pages":  failed_pages,
        "cost_usd":      round(cost, 4),
        "input_tokens":  combined_usage["input_tokens"],
        "output_tokens": combined_usage["output_tokens"],
        "google_pages":  combined_usage["google_pages"],
        "google_cost_usd": round(google_cost, 4),
        "output_files":  [
            _file_info(master_json_path),
            _file_info(csv_path),
            _file_info(master_path),
        ] + [
            _file_info(EXCEL_DIR / f"{pdf.stem.replace(' ', '_')}_extracted.xlsx")
            for pdf in selected
        ],
        "db_export":     db_status,
        "db_error":      db_error,
    }
    _write_run_log(run_entry)
    log.info("Run logged → logs/runs.log")

    log.info("=" * 62)
    log.info("Batch complete | %d pages | %d file(s)", pages_total, len(selected))
    if engine == "google":
        log.info(
            "Google Vision: %d pages ($%.4f)  |  Claude structuring: %d in + %d out tokens ($%.4f)  |  Total: $%.4f",
            combined_usage["google_pages"], google_cost,
            combined_usage["input_tokens"], combined_usage["output_tokens"],
            _actual_cost(combined_usage, model), cost,
        )
    else:
        log.info("API usage: %d in + %d out tokens = $%.4f  (model: %s)",
                 combined_usage["input_tokens"], combined_usage["output_tokens"], cost, model)
    log.info("=" * 62)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Timesheet extractor — Claude Vision / Google Vision")
    parser.add_argument("--model",     default=DEFAULT_MODEL,           help="Claude model ID")
    parser.add_argument("--engine",    default=None, choices=["claude", "google"],
                        help="Pre-select engine without prompting: 'claude' or 'google'. "
                             "If omitted, you will be asked interactively.")
    parser.add_argument("--output",    default="timesheet_output.xlsx", help="Master Excel filename")
    parser.add_argument("--delay",     type=float, default=0.5,         help="Seconds between API calls")
    parser.add_argument("--from-json", default=None, metavar="PATH",
                        help="Skip API calls — import a previously saved JSON file into the DB")
    args = parser.parse_args()

    if args.from_json:
        load_json_for_db_import(Path(args.from_json), model=args.model)
    else:
        main(
            model=args.model,
            output_excel=args.output,
            delay=args.delay,
            engine=args.engine,
        )
