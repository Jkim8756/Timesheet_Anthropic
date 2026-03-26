"""
Timesheet Batch Extractor
Processes timesheet images/PDFs via Claude Vision API -> Excel + optional DB export
- PDFs are split and processed one page at a time (best accuracy)
- Excel output: one sheet per page + relational flat sheet + summary
"""

import os
import io
import sys
import json
import base64
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# -- Logging -------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("extraction.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger(__name__)

# -- Config --------------------------------------------------------------------
SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf"}
MODEL = "claude-sonnet-4-20250514"
MAX_RETRIES = 3
RETRY_DELAY = 5

EXTRACTION_PROMPT = """
You are extracting data from a staff timesheet image (one page only).

Return ONLY a valid JSON object - no markdown, no explanation, no preamble.

Extract the following top-level fields:
- project (string)
- business_unit (string)
- date (string, as written)
- day_of_week (string)
- weather (string or null)

Extract "frontline_staff" as an array of objects, each with:
- job_task, title, employee_name, ein,
- scheduled_start, scheduled_end, scheduled_hours,
- actual_start, lunch_out, lunch_in, actual_end, actual_hours,
- signature (string or null),
- absent (true/false/null),
- schedule_changed (true/false/null),
- confidence (0.0-1.0 - lower if handwriting was hard to read)

Extract "management_staff" as an array with the same fields.

Extract "summary":
- attendees (integer or null)
- absent_count (integer or null)

If a field is illegible or missing, use null. Never guess - use null if uncertain.
"""

# -- Styles --------------------------------------------------------------------
def make_styles():
    thin = Side(style="thin")
    return {
        "bdr": Border(left=thin, right=thin, top=thin, bottom=thin),
        "hdr_fill": PatternFill("solid", start_color="1F3864", end_color="1F3864"),
        "alt_fill": PatternFill("solid", start_color="EEF2F7", end_color="EEF2F7"),
        "flag_fill": PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
        "fail_fill": PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0"),
        "partial_fill": PatternFill("solid", start_color="FFE5CC", end_color="FFE5CC"),
        "ok_fill": PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA"),
    }

STAFF_COLUMNS = [
    "Source File", "Page", "Date", "Day", "Business Unit", "Project",
    "Staff Type", "Job Task", "Title", "Employee Name", "EIN",
    "Sched Start", "Sched End", "Sched Hours",
    "Actual Start", "Lunch Out", "Lunch In", "Actual End", "Actual Hours",
    "Absent", "Schedule Changed", "Confidence", "Status"
]

STAFF_COL_WIDTHS = [
    22, 6, 12, 10, 14, 22, 12, 10, 14,
    22, 14, 11, 11, 11, 11, 11, 11, 11, 11,
    8, 16, 11, 16
]

# -- PDF page splitter ---------------------------------------------------------
def split_pdf_to_pages(pdf_path: Path) -> list[tuple[int, str]]:
    """Split a PDF into individual pages. Returns list of (page_number, base64_string)."""
    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        log.error("PDF splitting requires: pip install pypdf")
        raise

    reader = PdfReader(str(pdf_path))
    pages = []
    for i, page in enumerate(reader.pages, 1):
        writer = PdfWriter()
        writer.add_page(page)
        buf = io.BytesIO()
        writer.write(buf)
        b64 = base64.standard_b64encode(buf.getvalue()).decode("utf-8")
        pages.append((i, b64))
        log.info(f"  Prepared page {i}/{len(reader.pages)}")
    return pages


def encode_image(path: Path) -> tuple[str, str]:
    """Return (base64_data, media_type) for an image file."""
    ext = path.suffix.lower()
    media_map = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif",
        ".pdf": "application/pdf"
    }
    media_type = media_map.get(ext, "image/jpeg")
    with open(path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8"), media_type


# -- Claude API call -----------------------------------------------------------
def call_claude(client: anthropic.Anthropic, b64: str, media_type: str, label: str) -> Optional[dict]:
    """Send one page/image to Claude and return parsed JSON."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            content_block = {
                "type": "document" if media_type == "application/pdf" else "image",
                "source": {"type": "base64", "media_type": media_type, "data": b64}
            }
            response = client.messages.create(
                model=MODEL,
                max_tokens=8000,
                messages=[{
                    "role": "user",
                    "content": [content_block, {"type": "text", "text": EXTRACTION_PROMPT}]
                }]
            )
            raw = response.content[0].text.strip()
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]

            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                log.warning(f"[{label}] Response truncated, attempting partial recovery...")
                try:
                    salvaged = raw
                    last_brace = salvaged.rfind("},")
                    if last_brace == -1:
                        last_brace = salvaged.rfind("}")
                    if last_brace > 0:
                        salvaged = salvaged[:last_brace + 1]
                    salvaged += "]" * (salvaged.count("[") - salvaged.count("]"))
                    salvaged += "}" * (salvaged.count("{") - salvaged.count("}"))
                    result = json.loads(salvaged)
                    result["_partial"] = True
                    log.warning(f"[{label}] Partial recovery succeeded.")
                    return result
                except json.JSONDecodeError as e:
                    log.warning(f"[{label}] Partial recovery failed attempt {attempt}: {e}")

        except anthropic.RateLimitError:
            wait = RETRY_DELAY * attempt
            log.warning(f"[{label}] Rate limited - waiting {wait}s")
            time.sleep(wait)
        except Exception as e:
            log.error(f"[{label}] Error attempt {attempt}: {e}")
            time.sleep(RETRY_DELAY)

    return None


# -- Process a single file -----------------------------------------------------
def process_file(client: anthropic.Anthropic, file_path: Path, delay: float) -> list[dict]:
    """Process one file. PDFs are split per page. Returns list of page result dicts."""
    results = []
    ext = file_path.suffix.lower()

    if ext == ".pdf":
        log.info(f"  Splitting PDF into pages...")
        try:
            pages = split_pdf_to_pages(file_path)
        except Exception as e:
            log.error(f"  Failed to split PDF: {e}")
            return [{
                "_source_file": file_path.name, "_page": 1,
                "_extracted_at": datetime.now().isoformat(),
                "_status": "failed", "frontline_staff": [], "management_staff": []
            }]

        for page_num, b64 in pages:
            label = f"{file_path.name} p{page_num}"
            log.info(f"  Processing page {page_num}/{len(pages)}...")
            data = call_claude(client, b64, "application/pdf", label)
            result = _build_result(data, file_path.name, page_num)
            results.append(result)
            staff_count = (
                len(result.get("frontline_staff") or []) +
                len(result.get("management_staff") or [])
            )
            log.info(f"  >> {result['_status'].upper()} | page {page_num} | {staff_count} staff records")
            time.sleep(delay)
    else:
        b64, media_type = encode_image(file_path)
        data = call_claude(client, b64, media_type, file_path.name)
        result = _build_result(data, file_path.name, 1)
        results.append(result)
        staff_count = (
            len(result.get("frontline_staff") or []) +
            len(result.get("management_staff") or [])
        )
        log.info(f"  >> {result['_status'].upper()} | {staff_count} staff records")
        time.sleep(delay)

    return results


def _build_result(data: Optional[dict], source_file: str, page: int) -> dict:
    if data is None:
        return {
            "_source_file": source_file, "_page": page,
            "_extracted_at": datetime.now().isoformat(),
            "_status": "failed", "frontline_staff": [], "management_staff": []
        }
    data["_source_file"] = source_file
    data["_page"] = page
    data["_extracted_at"] = datetime.now().isoformat()
    data["_status"] = "partial" if data.pop("_partial", False) else "success"
    return data


# -- Excel helpers -------------------------------------------------------------
def style_header_row(ws, columns, styles, row=1):
    for col_idx, col_name in enumerate(columns, 1):
        c = ws.cell(row=row, column=col_idx, value=col_name)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = styles["hdr_fill"]
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = styles["bdr"]
    ws.row_dimensions[row].height = 28


def staff_row_values(result: dict, staff_type: str, s: dict) -> list:
    confidence = s.get("confidence", 1.0) or 1.0
    return [
        result.get("_source_file", ""),
        result.get("_page", ""),
        result.get("date", ""),
        result.get("day_of_week", ""),
        result.get("business_unit", ""),
        result.get("project", ""),
        staff_type,
        s.get("job_task", ""),
        s.get("title", ""),
        s.get("employee_name", ""),
        s.get("ein", ""),
        s.get("scheduled_start", ""),
        s.get("scheduled_end", ""),
        s.get("scheduled_hours", ""),
        s.get("actual_start", ""),
        s.get("lunch_out", ""),
        s.get("lunch_in", ""),
        s.get("actual_end", ""),
        s.get("actual_hours", ""),
        "Yes" if s.get("absent") is True else ("No" if s.get("absent") is False else ""),
        "Yes" if s.get("schedule_changed") is True else ("No" if s.get("schedule_changed") is False else ""),
        confidence,
        "Low Confidence" if confidence < 0.8 else "OK"
    ]


def write_staff_rows(ws, all_results, styles, start_row=2):
    row = start_row
    for result in all_results:
        if result["_status"] == "failed":
            c = ws.cell(row=row, column=1, value=result["_source_file"])
            c.fill = styles["fail_fill"]
            ws.cell(row=row, column=len(STAFF_COLUMNS), value="FAILED")
            row += 1
            continue

        all_staff = (
            [("Frontline", s) for s in (result.get("frontline_staff") or [])] +
            [("Management", s) for s in (result.get("management_staff") or [])]
        )
        page_fill = styles["partial_fill"] if result["_status"] == "partial" else None

        for staff_type, s in all_staff:
            confidence = s.get("confidence", 1.0) or 1.0
            if confidence < 0.8:
                row_fill = styles["flag_fill"]
            elif page_fill:
                row_fill = page_fill
            elif row % 2 == 0:
                row_fill = styles["alt_fill"]
            else:
                row_fill = None

            values = staff_row_values(result, staff_type, s)
            for col_idx, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = styles["bdr"]
                if row_fill:
                    c.fill = row_fill
            row += 1
    return row


# -- Excel Builder -------------------------------------------------------------
def build_excel(all_results: list[dict], output_path: Path):
    from collections import defaultdict
    wb = openpyxl.Workbook()
    styles = make_styles()

    # -- Sheet 1: Relational flat (all pages combined) ------------------------
    ws_flat = wb.active
    ws_flat.title = "All Records"
    style_header_row(ws_flat, STAFF_COLUMNS, styles)
    ws_flat.freeze_panes = "A2"
    write_staff_rows(ws_flat, all_results, styles)
    for i, w in enumerate(STAFF_COL_WIDTHS, 1):
        ws_flat.column_dimensions[get_column_letter(i)].width = w

    # -- One sheet per PDF page -----------------------------------------------
    grouped = defaultdict(list)
    for result in all_results:
        grouped[result["_source_file"]].append(result)

    for source_file, pages in grouped.items():
        for result in pages:
            page_num = result.get("_page", 1)
            raw_name = f"{Path(source_file).stem}_p{page_num}"
            sheet_name = raw_name[:31]
            for ch in r'/\*?[]:|':
                sheet_name = sheet_name.replace(ch, "")

            ws_page = wb.create_sheet(title=sheet_name)

            # Meta block
            meta_rows = [
                ("Project", result.get("project", "")),
                ("Business Unit", result.get("business_unit", "")),
                ("Date", result.get("date", "")),
                ("Day", result.get("day_of_week", "")),
                ("Weather", result.get("weather", "")),
                ("Page", page_num),
                ("Status", result.get("_status", "").upper()),
            ]
            for r_idx, (label, val) in enumerate(meta_rows, 1):
                c = ws_page.cell(row=r_idx, column=1, value=label)
                c.font = Font(bold=True, name="Calibri", size=10)
                ws_page.cell(row=r_idx, column=2, value=val).font = Font(name="Calibri", size=10)
            ws_page.column_dimensions["A"].width = 16
            ws_page.column_dimensions["B"].width = 24

            # Staff table at row 10
            style_header_row(ws_page, STAFF_COLUMNS, styles, row=10)
            ws_page.freeze_panes = "A11"
            last_row = write_staff_rows(ws_page, [result], styles, start_row=11)
            for i, w in enumerate(STAFF_COL_WIDTHS, 1):
                ws_page.column_dimensions[get_column_letter(i)].width = w

            # Summary block
            summary = result.get("summary") or {}
            ws_page.cell(row=last_row + 1, column=1, value="Attendees").font = Font(bold=True, name="Calibri")
            ws_page.cell(row=last_row + 1, column=2, value=summary.get("attendees"))
            ws_page.cell(row=last_row + 2, column=1, value="Absent").font = Font(bold=True, name="Calibri")
            ws_page.cell(row=last_row + 2, column=2, value=summary.get("absent_count"))

    # -- Summary sheet --------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Batch Extraction Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, name="Calibri", color="1F3864")

    total_pages = len(all_results)
    success = sum(1 for r in all_results if r["_status"] == "success")
    partial = sum(1 for r in all_results if r["_status"] == "partial")
    failed = sum(1 for r in all_results if r["_status"] == "failed")
    total_staff = sum(
        len(r.get("frontline_staff") or []) + len(r.get("management_staff") or [])
        for r in all_results
    )
    low_conf = sum(
        1 for r in all_results
        for s in (r.get("frontline_staff") or []) + (r.get("management_staff") or [])
        if (s.get("confidence") or 1.0) < 0.8
    )

    summary_data = [
        ("", ""),
        ("Total pages processed", total_pages),
        ("Successful", success),
        ("Partial (truncated)", partial),
        ("Failed", failed),
        ("Total staff records", total_staff),
        ("Low confidence rows", low_conf),
        ("", ""),
        ("Extracted at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Color Legend", ""),
        ("Yellow row", "Low confidence (<80%) - review recommended"),
        ("Orange row", "Partial result - response was truncated"),
        ("Red row", "Failed - no data extracted"),
        ("Green row", "Success"),
    ]

    legend_fills = {
        "Yellow row": styles["flag_fill"],
        "Orange row": styles["partial_fill"],
        "Red row": styles["fail_fill"],
        "Green row": styles["ok_fill"],
    }

    for r_idx, (label, val) in enumerate(summary_data, 2):
        c = ws_sum.cell(row=r_idx, column=1, value=label)
        c.font = Font(bold=True, name="Calibri", size=10)
        v = ws_sum.cell(row=r_idx, column=2, value=val)
        v.font = Font(name="Calibri", size=10)
        if label in legend_fills:
            c.fill = legend_fills[label]
            v.fill = legend_fills[label]

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 45

    # Move Summary to front
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))

    wb.save(output_path)
    log.info(f"Excel saved -> {output_path}")


# -- Database Export -----------------------------------------------------------
def export_to_db(all_results: list[dict], db_url: str):
    try:
        from sqlalchemy import create_engine, text
        import pandas as pd
    except ImportError:
        log.error("DB export requires: pip install psycopg2-binary sqlalchemy pandas")
        return

    rows = []
    for result in all_results:
        if result["_status"] == "failed":
            continue
        for staff_type, staff_list in [
            ("frontline", result.get("frontline_staff") or []),
            ("management", result.get("management_staff") or [])
        ]:
            for s in staff_list:
                rows.append({
                    "source_file": result.get("_source_file"),
                    "page": result.get("_page"),
                    "extracted_at": result.get("_extracted_at"),
                    "extraction_status": result.get("_status"),
                    "project": result.get("project"),
                    "business_unit": result.get("business_unit"),
                    "work_date": result.get("date"),
                    "day_of_week": result.get("day_of_week"),
                    "staff_type": staff_type,
                    "job_task": s.get("job_task"),
                    "title": s.get("title"),
                    "employee_name": s.get("employee_name"),
                    "ein": s.get("ein"),
                    "scheduled_start": s.get("scheduled_start"),
                    "scheduled_end": s.get("scheduled_end"),
                    "scheduled_hours": s.get("scheduled_hours"),
                    "actual_start": s.get("actual_start"),
                    "lunch_out": s.get("lunch_out"),
                    "lunch_in": s.get("lunch_in"),
                    "actual_end": s.get("actual_end"),
                    "actual_hours": s.get("actual_hours"),
                    "absent": s.get("absent"),
                    "schedule_changed": s.get("schedule_changed"),
                    "confidence": s.get("confidence"),
                })

    df = pd.DataFrame(rows)
    engine = create_engine(db_url)
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS timesheet_staging (
                id SERIAL PRIMARY KEY,
                source_file VARCHAR(255),
                page INTEGER,
                extracted_at TIMESTAMP,
                extraction_status VARCHAR(20),
                project VARCHAR(255),
                business_unit VARCHAR(100),
                work_date VARCHAR(50),
                day_of_week VARCHAR(20),
                staff_type VARCHAR(20),
                job_task VARCHAR(10),
                title VARCHAR(100),
                employee_name VARCHAR(150),
                ein VARCHAR(30),
                scheduled_start VARCHAR(20),
                scheduled_end VARCHAR(20),
                scheduled_hours VARCHAR(10),
                actual_start VARCHAR(20),
                lunch_out VARCHAR(20),
                lunch_in VARCHAR(20),
                actual_end VARCHAR(20),
                actual_hours VARCHAR(10),
                absent BOOLEAN,
                schedule_changed BOOLEAN,
                confidence NUMERIC(3,2),
                reviewed BOOLEAN DEFAULT FALSE
            )
        """))
        conn.commit()
    df.to_sql("timesheet_staging", engine, if_exists="append", index=False)
    log.info(f"Exported {len(df)} rows to database staging table.")


# -- Main Runner ---------------------------------------------------------------
def run_batch(
    input_folder: str,
    output_excel: str = "timesheet_output.xlsx",
    results_json: str = "results.json",
    db_url: Optional[str] = None,
    delay_between_requests: float = 0.5
):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError("Set ANTHROPIC_API_KEY in your .env file.")

    # Check Excel is not open before spending any credits
    if Path(output_excel).exists():
        try:
            test = open(output_excel, "a")
            test.close()
        except IOError:
            log.error(f"ERROR: '{output_excel}' is open in Excel. Please close it and run again.")
            return

    client = anthropic.Anthropic(api_key=api_key)
    input_path = Path(input_folder)
    files = sorted([f for f in input_path.iterdir() if f.suffix.lower() in SUPPORTED_EXTENSIONS])

    if not files:
        log.warning(f"No supported files found in: {input_folder}")
        return

    log.info(f"Found {len(files)} file(s) to process.")
    all_results = []

    for i, file_path in enumerate(files, 1):
        log.info(f"[{i}/{len(files)}] Processing: {file_path.name}")
        page_results = process_file(client, file_path, delay_between_requests)
        all_results.extend(page_results)

        # Incremental save after each file (crash recovery)
        with open(results_json, "w", encoding="utf-8") as f:
            json.dump(all_results, f, indent=2)

    build_excel(all_results, Path(output_excel))

    if db_url:
        export_to_db(all_results, db_url)

    success = sum(1 for r in all_results if r["_status"] == "success")
    partial = sum(1 for r in all_results if r["_status"] == "partial")
    failed = sum(1 for r in all_results if r["_status"] == "failed")
    log.info("=" * 50)
    log.info(f"Batch complete: {success} OK, {partial} partial, {failed} failed")
    log.info(f"Total pages processed: {len(all_results)}")
    log.info(f"Excel output: {output_excel}")
    log.info(f"JSON backup:  {results_json}")
    log.info("=" * 50)


# -- Entry Point ---------------------------------------------------------------
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Batch timesheet extractor using Claude Vision")
    parser.add_argument("input_folder", help="Folder containing timesheet images/PDFs")
    parser.add_argument("--output", default="timesheet_output.xlsx", help="Output Excel filename")
    parser.add_argument("--json", default="results.json", help="JSON backup filename")
    parser.add_argument("--db", default=None, help="PostgreSQL URL for DB export (optional)")
    parser.add_argument("--delay", type=float, default=0.5, help="Seconds between API calls (default: 0.5)")
    args = parser.parse_args()

    run_batch(
        input_folder=args.input_folder,
        output_excel=args.output,
        results_json=args.json,
        db_url=args.db,
        delay_between_requests=args.delay
    )