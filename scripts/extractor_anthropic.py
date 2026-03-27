"""
Timesheet Batch Extractor
Processes timesheet PDFs (zip-of-jpegs format) via Claude Vision API
-> per-source Excel files + master Excel + optional SQL export

Usage:
    python extractor_anthropic.py timesheets_pdf
    python extractor_anthropic.py timesheets_pdf --model claude-opus-4-20250514
    python extractor_anthropic.py timesheets_pdf --output my_output.xlsx
    python extractor_anthropic.py timesheets_pdf --db postgresql://user:pass@host/db
"""

import os
import sys
import json
import base64
import zipfile
import time
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Optional
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()  # try current directory first
load_dotenv(Path(__file__).parent.parent / "config" / ".env")  # fallback: config/.env

# -- Config --------------------------------------------------------------------
SUPPORTED_EXTENSIONS = {".pdf"}  # zip-of-jpegs format

# Token cost estimates per page (used for pre-flight estimate only)
INPUT_TOKENS_PER_PAGE  = 1600
OUTPUT_TOKENS_PER_PAGE = 600

# Pricing per million tokens (update if model pricing changes)
MODEL_PRICING = {
    # model substring -> (input $/Mtok, output $/Mtok)
    "claude-opus-4":    (15.00, 75.00),
    "claude-sonnet-4":  (3.00,  15.00),
    "claude-haiku-4":   (0.80,  4.00),
    "claude-haiku-3":   (0.25,  1.25),
    "default":          (3.00,  15.00),
}

DEFAULT_MODEL = "claude-sonnet-4-20250514"
MAX_RETRIES   = 3
RETRY_DELAY   = 5

EXTRACTION_PROMPT = """
You are extracting data from a staff timesheet image (one page only).

Return ONLY a valid JSON object - no markdown, no explanation, no preamble.

Extract the following top-level fields:
- project (string)
- business_unit (string)
- date (string, as written)
- day_of_week (string)
- weather (string or null)

Extract "frontline_staff" as an array of objects, each with:
- job_task, title, employee_name, ein,
- scheduled_start, scheduled_end, scheduled_hours,
- actual_start, lunch_out, lunch_in, actual_end, actual_hours,
- signature (string or null),
- absent (true/false/null),
- schedule_changed (true/false/null),
- confidence (0.0-1.0 - lower if handwriting was hard to read)

Extract "management_staff" as an array with the same fields.

Extract "summary":
- attendees (integer or null)
- absent_count (integer or null)

If a field is illegible or missing, use null. Never guess - use null if uncertain.
"""

STAFF_COLUMNS = [
    "Source File", "Page", "Date", "Day", "Business Unit", "Project",
    "Staff Type", "Job Task", "Title", "Employee Name", "EIN",
    "Sched Start", "Sched End", "Sched Hours",
    "Actual Start", "Lunch Out", "Lunch In", "Actual End", "Actual Hours",
    "Absent", "Schedule Changed", "Confidence", "Status"
]

STAFF_COL_WIDTHS = [
    22, 6, 12, 10, 14, 22, 12, 10, 14,
    22, 14, 11, 11, 11, 11, 11, 11, 11, 11,
    8, 16, 11, 16
]


# -- Cost helpers --------------------------------------------------------------
def get_pricing(model: str) -> tuple[float, float]:
    for key, pricing in MODEL_PRICING.items():
        if key in model:
            return pricing
    return MODEL_PRICING["default"]


def estimate_cost(total_pages: int, model: str) -> float:
    inp, out = get_pricing(model)
    return (total_pages * INPUT_TOKENS_PER_PAGE / 1_000_000 * inp +
            total_pages * OUTPUT_TOKENS_PER_PAGE / 1_000_000 * out)


def actual_cost(usage: dict, model: str) -> float:
    inp, out = get_pricing(model)
    return (usage.get("input_tokens", 0) / 1_000_000 * inp +
            usage.get("output_tokens", 0) / 1_000_000 * out)


# -- Source file reader --------------------------------------------------------
def get_pages(pdf_path: Path) -> list[dict]:
    """Auto-detect format: zip-of-jpegs (manifest.json) or real PDF."""
    # Try zip-of-jpegs first
    if zipfile.is_zipfile(pdf_path):
        pages = []
        with zipfile.ZipFile(pdf_path) as zf:
            manifest = json.loads(zf.read("manifest.json"))
            for p in manifest["pages"]:
                img_data = zf.read(p["image"]["path"])
                pages.append({
                    "page_number": p["page_number"],
                    "image_b64": base64.standard_b64encode(img_data).decode(),
                    "media_type": p["image"]["media_type"],
                })
        return pages

    # Real PDF - split per page via pypdf
    import io
    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        raise ImportError("Real PDF support requires: pip install pypdf")

    reader = PdfReader(str(pdf_path))
    pages = []
    for i, page in enumerate(reader.pages, 1):
        writer = PdfWriter()
        writer.add_page(page)
        buf = io.BytesIO()
        writer.write(buf)
        b64 = base64.standard_b64encode(buf.getvalue()).decode()
        pages.append({
            "page_number": i,
            "image_b64": b64,
            "media_type": "application/pdf",
        })
    return pages


# -- Pre-flight file selector --------------------------------------------------
def select_files(input_path: Path, model: str) -> list[Path]:
    seen = set()
    candidates = []
    for f in sorted(input_path.iterdir()):
        if f.suffix.lower() == ".pdf" and f.name.lower() not in seen:
            seen.add(f.name.lower())
            candidates.append(f)
    if not candidates:
        print(f"No .pdf files found in: {input_path}")
        sys.exit(0)

    # Peek page counts
    page_counts = {}
    for f in candidates:
        try:
            if zipfile.is_zipfile(f):
                with zipfile.ZipFile(f) as zf:
                    page_counts[f] = json.loads(zf.read("manifest.json"))["num_pages"]
            else:
                from pypdf import PdfReader
                page_counts[f] = len(PdfReader(str(f)).pages)
        except Exception:
            page_counts[f] = "?"

    print("\n" + "=" * 62)
    print("  TIMESHEET EXTRACTOR — FILE SELECTION")
    print("=" * 62)
    print(f"  Model : {model}")
    print()
    for i, f in enumerate(candidates, 1):
        kb = f.stat().st_size / 1024
        print(f"  [{i}] {f.name:<42} {page_counts[f]:>3} pages  ({kb:.0f} KB)")

    print("\nSelect files to process (e.g. 1,2  or  all  or  q to quit):")
    while True:
        raw = input("  > ").strip().lower()
        if raw == "q":
            print("Exiting.")
            sys.exit(0)
        if raw == "all":
            selected = candidates
            break
        try:
            indices = [int(x.strip()) - 1 for x in raw.split(",")]
            selected = [candidates[i] for i in indices if 0 <= i < len(candidates)]
            if selected:
                break
            print("  No valid selections. Try again.")
        except ValueError:
            print("  Invalid input.")

    total_pages = sum(page_counts[f] for f in selected if isinstance(page_counts[f], int))
    est = estimate_cost(total_pages, model)

    print("\n" + "-" * 62)
    print("  FILES SELECTED:")
    for f in selected:
        print(f"    • {f.name}  ({page_counts[f]} pages)")
    print(f"\n  Total pages    : {total_pages}")
    print(f"  Model          : {model}")
    print(f"  Estimated cost : ${est:.4f} USD")
    print("-" * 62)
    print("\nProceed? (y/n):")
    if input("  > ").strip().lower() != "y":
        print("Cancelled.")
        sys.exit(0)

    return selected


# -- Claude API call -----------------------------------------------------------
def call_claude(client: anthropic.Anthropic, page: dict, label: str) -> Optional[dict]:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.messages.create(
                model=client._model,
                max_tokens=8000,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "document" if page["media_type"] == "application/pdf" else "image",
                            "source": {
                                "type": "base64",
                                "media_type": page["media_type"],
                                "data": page["image_b64"],
                            }
                        },
                        {"type": "text", "text": EXTRACTION_PROMPT}
                    ]
                }]
            )
            raw = response.content[0].text.strip()
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
            data = json.loads(raw.strip())
            usage = {
                "input_tokens":  response.usage.input_tokens,
                "output_tokens": response.usage.output_tokens,
            }
            return data, usage

        except json.JSONDecodeError:
            print(f"    ⚠  [{label}] JSON parse failed attempt {attempt}")
            if attempt == MAX_RETRIES:
                return None, {}
        except anthropic.RateLimitError:
            wait = RETRY_DELAY * attempt
            print(f"    ⚠  [{label}] Rate limited — waiting {wait}s")
            time.sleep(wait)
        except Exception as e:
            print(f"    ⚠  [{label}] Error attempt {attempt}: {e}")
            time.sleep(RETRY_DELAY)

    return None, {}


def _build_result(data: Optional[dict], source_file: str, page: int) -> dict:
    if data is None:
        return {
            "_source_file": source_file, "_page": page,
            "_extracted_at": datetime.now().isoformat(),
            "_status": "failed", "frontline_staff": [], "management_staff": []
        }
    data["_source_file"] = source_file
    data["_page"] = page
    data["_extracted_at"] = datetime.now().isoformat()
    data["_status"] = "success"
    return data


# -- Excel styles --------------------------------------------------------------
def make_styles():
    thin = Side(style="thin")
    return {
        "bdr":          Border(left=thin, right=thin, top=thin, bottom=thin),
        "hdr_fill":     PatternFill("solid", start_color="1F3864"),
        "alt_fill":     PatternFill("solid", start_color="EEF2F7"),
        "flag_fill":    PatternFill("solid", start_color="FFF2CC"),
        "fail_fill":    PatternFill("solid", start_color="FFE0E0"),
        "partial_fill": PatternFill("solid", start_color="FFE5CC"),
        "ok_fill":      PatternFill("solid", start_color="E2EFDA"),
    }


def style_header_row(ws, columns, styles, row=1):
    for col_idx, col_name in enumerate(columns, 1):
        c = ws.cell(row=row, column=col_idx, value=col_name)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = styles["hdr_fill"]
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = styles["bdr"]
    ws.row_dimensions[row].height = 28


def staff_row_values(result: dict, staff_type: str, s: dict) -> list:
    confidence = s.get("confidence", 1.0) or 1.0
    return [
        result.get("_source_file", ""),
        result.get("_page", ""),
        result.get("date", ""),
        result.get("day_of_week", ""),
        result.get("business_unit", ""),
        result.get("project", ""),
        staff_type,
        s.get("job_task", ""),
        s.get("title", ""),
        s.get("employee_name", ""),
        s.get("ein", ""),
        s.get("scheduled_start", ""),
        s.get("scheduled_end", ""),
        s.get("scheduled_hours", ""),
        s.get("actual_start", ""),
        s.get("lunch_out", ""),
        s.get("lunch_in", ""),
        s.get("actual_end", ""),
        s.get("actual_hours", ""),
        "Yes" if s.get("absent") is True else ("No" if s.get("absent") is False else ""),
        "Yes" if s.get("schedule_changed") is True else ("No" if s.get("schedule_changed") is False else ""),
        confidence,
        "Low Confidence" if confidence < 0.8 else "OK"
    ]


def write_staff_rows(ws, all_results, styles, start_row=2):
    row = start_row
    for result in all_results:
        if result["_status"] == "failed":
            c = ws.cell(row=row, column=1, value=result["_source_file"])
            c.fill = styles["fail_fill"]
            ws.cell(row=row, column=len(STAFF_COLUMNS), value="FAILED")
            row += 1
            continue

        all_staff = (
            [("Frontline", s) for s in (result.get("frontline_staff") or [])] +
            [("Management", s) for s in (result.get("management_staff") or [])]
        )

        for staff_type, s in all_staff:
            confidence = s.get("confidence", 1.0) or 1.0
            if confidence < 0.8:
                row_fill = styles["flag_fill"]
            elif row % 2 == 0:
                row_fill = styles["alt_fill"]
            else:
                row_fill = None

            values = staff_row_values(result, staff_type, s)
            for col_idx, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = styles["bdr"]
                if row_fill:
                    c.fill = row_fill
            row += 1
    return row


# -- Excel builder (per-source file) -------------------------------------------
def build_excel(all_results: list[dict], output_path: Path,
                model: str, usage: dict):
    wb = openpyxl.Workbook()
    styles = make_styles()

    # Sheet 1: All Records (flat relational)
    ws_flat = wb.active
    ws_flat.title = "All Records"
    style_header_row(ws_flat, STAFF_COLUMNS, styles)
    ws_flat.freeze_panes = "A2"
    write_staff_rows(ws_flat, all_results, styles)
    for i, w in enumerate(STAFF_COL_WIDTHS, 1):
        ws_flat.column_dimensions[get_column_letter(i)].width = w

    # Per-page sheets
    grouped = defaultdict(list)
    for result in all_results:
        grouped[result["_source_file"]].append(result)

    for source_file, pages in grouped.items():
        for result in pages:
            page_num = result.get("_page", 1)
            raw_name = f"{Path(source_file).stem}_p{page_num}"
            sheet_name = raw_name[:31]
            for ch in r'/\*?[]:|':
                sheet_name = sheet_name.replace(ch, "")

            ws_p = wb.create_sheet(title=sheet_name)

            meta = [
                ("Project",       result.get("project", "")),
                ("Business Unit", result.get("business_unit", "")),
                ("Date",          result.get("date", "")),
                ("Day",           result.get("day_of_week", "")),
                ("Weather",       result.get("weather", "")),
                ("Page",          page_num),
                ("Status",        result.get("_status", "").upper()),
            ]
            for r_idx, (label, val) in enumerate(meta, 1):
                ws_p.cell(row=r_idx, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
                ws_p.cell(row=r_idx, column=2, value=val).font = Font(name="Calibri", size=10)
            ws_p.column_dimensions["A"].width = 16
            ws_p.column_dimensions["B"].width = 24

            style_header_row(ws_p, STAFF_COLUMNS, styles, row=10)
            ws_p.freeze_panes = "A11"
            last_row = write_staff_rows(ws_p, [result], styles, start_row=11)
            for i, w in enumerate(STAFF_COL_WIDTHS, 1):
                ws_p.column_dimensions[get_column_letter(i)].width = w

            summary = result.get("summary") or {}
            ws_p.cell(row=last_row + 1, column=1, value="Attendees").font = Font(bold=True, name="Calibri")
            ws_p.cell(row=last_row + 1, column=2, value=summary.get("attendees"))
            ws_p.cell(row=last_row + 2, column=1, value="Absent").font = Font(bold=True, name="Calibri")
            ws_p.cell(row=last_row + 2, column=2, value=summary.get("absent_count"))

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Batch Extraction Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, name="Calibri", color="1F3864")

    total_pages  = len(all_results)
    success      = sum(1 for r in all_results if r["_status"] == "success")
    failed       = sum(1 for r in all_results if r["_status"] == "failed")
    total_staff  = sum(
        len(r.get("frontline_staff") or []) + len(r.get("management_staff") or [])
        for r in all_results
    )
    low_conf = sum(
        1 for r in all_results
        for s in (r.get("frontline_staff") or []) + (r.get("management_staff") or [])
        if (s.get("confidence") or 1.0) < 0.8
    )
    cost = actual_cost(usage, model)

    summary_data = [
        ("", ""),
        ("Model used",            model),
        ("Total pages processed", total_pages),
        ("Successful",            success),
        ("Failed",                failed),
        ("Total staff records",   total_staff),
        ("Low confidence rows",   low_conf),
        ("Input tokens",          usage.get("input_tokens", "N/A")),
        ("Output tokens",         usage.get("output_tokens", "N/A")),
        ("Actual cost (USD)",     f"${cost:.4f}"),
        ("",                      ""),
        ("Extracted at",          datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("",                      ""),
        ("Color Legend",          ""),
        ("Yellow row",            "Low confidence (<80%) — review recommended"),
        ("Red row",               "Failed — no data extracted"),
        ("Green row",             "Success"),
    ]

    legend_fills = {
        "Yellow row": styles["flag_fill"],
        "Red row":    styles["fail_fill"],
        "Green row":  styles["ok_fill"],
    }

    for r_idx, (label, val) in enumerate(summary_data, 2):
        c = ws_sum.cell(row=r_idx, column=1, value=label)
        c.font = Font(bold=True, name="Calibri", size=10)
        v = ws_sum.cell(row=r_idx, column=2, value=val)
        v.font = Font(name="Calibri", size=10)
        if label in legend_fills:
            c.fill = legend_fills[label]
            v.fill = legend_fills[label]

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 45
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))

    wb.save(output_path)
    print(f"  ✅ Excel saved → {output_path.name}")


# -- Master Excel (all sources merged) ----------------------------------------
def build_master_excel(all_results: list[dict], output_path: Path,
                       model: str, usage: dict, source_files: list[str]):
    wb = openpyxl.Workbook()
    styles = make_styles()

    ws_flat = wb.active
    ws_flat.title = "All Records"
    style_header_row(ws_flat, STAFF_COLUMNS, styles)
    ws_flat.freeze_panes = "A2"
    write_staff_rows(ws_flat, all_results, styles)
    for i, w in enumerate(STAFF_COL_WIDTHS, 1):
        ws_flat.column_dimensions[get_column_letter(i)].width = w

    # Run Summary
    ws_sum = wb.create_sheet("Run Summary")
    ws_sum["A1"] = "Pipeline Run Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, name="Calibri", color="1F3864")

    total_staff = sum(
        len(r.get("frontline_staff") or []) + len(r.get("management_staff") or [])
        for r in all_results
    )
    cost = actual_cost(usage, model)

    summary_data = [
        ("", ""),
        ("Model used",          model),
        ("Source files",        ", ".join(source_files)),
        ("Total pages",         len(all_results)),
        ("Total staff records", total_staff),
        ("Input tokens",        usage.get("input_tokens", "N/A")),
        ("Output tokens",       usage.get("output_tokens", "N/A")),
        ("Actual cost (USD)",   f"${cost:.4f}"),
        ("Generated at",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for r_idx, (label, val) in enumerate(summary_data, 2):
        ws_sum.cell(row=r_idx, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
        ws_sum.cell(row=r_idx, column=2, value=val).font = Font(name="Calibri", size=10)

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 60
    wb.move_sheet("Run Summary", offset=-(len(wb.sheetnames) - 1))

    wb.save(output_path)
    print(f"  ✅ Master Excel saved → {output_path.name}")


# -- SQL / DB export -----------------------------------------------------------
def export_to_db(all_results: list[dict], db_url: str, model: str):
    try:
        from sqlalchemy import create_engine, text
        import pandas as pd
    except ImportError:
        print("DB export requires: pip install psycopg2-binary sqlalchemy pandas")
        return

    rows = []
    for result in all_results:
        if result["_status"] == "failed":
            continue
        for staff_type, staff_list in [
            ("Frontline",  result.get("frontline_staff")  or []),
            ("Management", result.get("management_staff") or []),
        ]:
            for s in staff_list:
                rows.append({
                    "source_file":       result.get("_source_file"),
                    "page":              result.get("_page"),
                    "extracted_at":      result.get("_extracted_at"),
                    "extraction_status": result.get("_status"),
                    "model_used":        model,
                    "project":           result.get("project"),
                    "business_unit":     result.get("business_unit"),
                    "work_date":         result.get("date"),
                    "day_of_week":       result.get("day_of_week"),
                    "staff_type":        staff_type,
                    "job_task":          s.get("job_task"),
                    "title":             s.get("title"),
                    "employee_name":     s.get("employee_name"),
                    "ein":               s.get("ein"),
                    "scheduled_start":   s.get("scheduled_start"),
                    "scheduled_end":     s.get("scheduled_end"),
                    "scheduled_hours":   s.get("scheduled_hours"),
                    "actual_start":      s.get("actual_start"),
                    "lunch_out":         s.get("lunch_out"),
                    "lunch_in":          s.get("lunch_in"),
                    "actual_end":        s.get("actual_end"),
                    "actual_hours":      s.get("actual_hours"),
                    "absent":            s.get("absent"),
                    "schedule_changed":  s.get("schedule_changed"),
                    "confidence":        s.get("confidence"),
                })

    if not rows:
        print("  No rows to export to DB.")
        return

    df = pd.DataFrame(rows)
    engine = create_engine(db_url)
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS timesheet_staging (
                id                 SERIAL PRIMARY KEY,
                source_file        VARCHAR(255),
                page               INTEGER,
                extracted_at       TIMESTAMP,
                extraction_status  VARCHAR(20),
                model_used         VARCHAR(100),
                project            VARCHAR(255),
                business_unit      VARCHAR(100),
                work_date          VARCHAR(50),
                day_of_week        VARCHAR(20),
                staff_type         VARCHAR(20),
                job_task           VARCHAR(10),
                title              VARCHAR(100),
                employee_name      VARCHAR(150),
                ein                VARCHAR(30),
                scheduled_start    VARCHAR(20),
                scheduled_end      VARCHAR(20),
                scheduled_hours    VARCHAR(10),
                actual_start       VARCHAR(20),
                lunch_out          VARCHAR(20),
                lunch_in           VARCHAR(20),
                actual_end         VARCHAR(20),
                actual_hours       VARCHAR(10),
                absent             BOOLEAN,
                schedule_changed   BOOLEAN,
                confidence         NUMERIC(3,2),
                reviewed           BOOLEAN DEFAULT FALSE
            )
        """))
        conn.commit()

    df.to_sql("timesheet_staging", engine, if_exists="append", index=False)
    print(f"  ✅ Exported {len(df)} rows → timesheet_staging")


# -- Main runner ---------------------------------------------------------------
def run_batch(input_folder: str, output_dir: str, output_excel: str, model: str,
              db_url: Optional[str], delay: float):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError("Set ANTHROPIC_API_KEY in your .env file.")

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    master_path = out_dir / output_excel

    # Check Excel not open before spending credits
    if master_path.exists():
        try:
            open(master_path, "a").close()
        except IOError:
            print(f"ERROR: '{master_path}' is open in Excel. Close it and run again.")
            return

    client = anthropic.Anthropic(api_key=api_key)
    client._model = model  # attach model to client for convenience

    input_path = Path(input_folder)
    selected = select_files(input_path, model)

    all_master_results = []
    all_source_names   = []
    combined_usage     = {"input_tokens": 0, "output_tokens": 0}

    for file_path in selected:
        print(f"\n📄 {file_path.name}")
        pages = get_pages(file_path)
        file_results = []
        file_usage   = {"input_tokens": 0, "output_tokens": 0}

        for page in pages:
            pnum  = page["page_number"]
            label = f"{file_path.name} p{pnum}"
            print(f"  Page {pnum}/{len(pages)} ...", end=" ", flush=True)

            data, usage = call_claude(client, page, label)
            result = _build_result(data, file_path.name, pnum)
            file_results.append(result)

            file_usage["input_tokens"]  += usage.get("input_tokens", 0)
            file_usage["output_tokens"] += usage.get("output_tokens", 0)

            n_staff = (len(result.get("frontline_staff") or []) +
                       len(result.get("management_staff") or []))
            status = result["_status"].upper()
            print(f"{status} ({n_staff} staff)")

            time.sleep(delay)

        # Per-source Excel → output dir
        stem = Path(file_path.stem).name.replace(" ", "_")
        per_source_path = out_dir / f"{stem}_extracted.xlsx"
        build_excel(file_results, per_source_path, model, file_usage)

        all_master_results.extend(file_results)
        all_source_names.append(file_path.name)
        combined_usage["input_tokens"]  += file_usage["input_tokens"]
        combined_usage["output_tokens"] += file_usage["output_tokens"]

    # Master Excel → output dir
    build_master_excel(all_master_results, master_path, model,
                       combined_usage, all_source_names)

    # DB export
    if db_url:
        print("\n🗄  Exporting to database...")
        export_to_db(all_master_results, db_url, model)

    cost = actual_cost(combined_usage, model)
    print(f"\n{'=' * 62}")
    print(f"  Batch complete | {len(all_master_results)} pages | {len(all_source_names)} files")
    print(f"  💰 Actual cost : ${cost:.4f} USD")
    print(f"     Input  tokens: {combined_usage['input_tokens']:,}")
    print(f"     Output tokens: {combined_usage['output_tokens']:,}")
    print(f"{'=' * 62}")


# -- Entry point ---------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Batch timesheet extractor using Claude Vision")
    parser.add_argument("input_folder",   help="Folder containing timesheet .pdf files")
    parser.add_argument("--output-dir",   default="data/2.output",         help="Directory for all Excel outputs (default: data/2.output)")
    parser.add_argument("--output",       default="timesheet_output.xlsx", help="Master Excel filename (default: timesheet_output.xlsx)")
    parser.add_argument("--model",        default=DEFAULT_MODEL,           help="Claude model to use")
    parser.add_argument("--db",           default=None,                    help="PostgreSQL URL for DB export")
    parser.add_argument("--delay",        type=float, default=0.5,         help="Seconds between API calls (default: 0.5)")
    args = parser.parse_args()

    run_batch(
        input_folder=args.input_folder,
        output_dir=args.output_dir,
        output_excel=args.output,
        model=args.model,
        db_url=args.db,
        delay=args.delay,
    )