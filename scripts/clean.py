"""
Timesheet Cleaner
=================
No API calls. Loads an extracted Excel file, cleans All Records,
and saves a new *_cleaned.xlsx to data/3.archive.

Usage:
    python clean.py                          # interactive file picker
    python clean.py --input path/to/file.xlsx
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Output schema (matches secai_cleaned.xlsx "Cleaned" sheet) ────────────────
CLEAN_COLS = [
    "Source File", "Page", "Date", "Day", "Business Unit", "Project",
    "EIN", "Employee Name", "Title", "Recorded Hours",
    "Staff Type", "Job Task",
    "Sched Start", "Sched End", "Sched Hours",
    "Actual Start", "Lunch Out", "Lunch In", "Actual End", "Actual Hours",
    "Absent", "Schedule Changed", "Confidence", "Status"
]

# Canonical Business Unit — strip common OCR noise
BU_PATTERN = re.compile(r'[^0-9.]')

# Canonical day names (EN + ES)
DAY_MAP = {
    "lunes": "Monday", "martes": "Tuesday", "miércoles": "Wednesday",
    "miercoles": "Wednesday", "jueves": "Thursday", "viernes": "Friday",
    "sábado": "Saturday", "sabado": "Saturday", "domingo": "Sunday",
    "monday": "Monday", "tuesday": "Tuesday", "wednesday": "Wednesday",
    "thursday": "Thursday", "friday": "Friday", "saturday": "Saturday",
    "sunday": "Sunday",
    # abbreviations
    "mon": "Monday", "tue": "Tuesday", "wed": "Wednesday",
    "thu": "Thursday", "fri": "Friday", "sat": "Saturday", "sun": "Sunday",
    "lun": "Monday", "mar": "Tuesday", "mié": "Wednesday", "jue": "Thursday",
    "vie": "Friday", "sáb": "Saturday", "dom": "Sunday",
}

# Time normalizer: turn "7AM", "7 AM", "7:00AM", "7 am" → "7:00AM"
_TIME_RE = re.compile(
    r'^\s*(\d{1,2})(?::(\d{2}))?\s*(am|pm|AM|PM|Am|Pm)?\s*$', re.IGNORECASE
)

def normalize_time(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "-", "--", "—"):
        return None
    m = _TIME_RE.match(s)
    if m:
        h, mins, meridiem = m.groups()
        mins = mins or "00"
        meridiem = (meridiem or "").upper()
        if not meridiem:
            # guess AM/PM from context — leave blank if unknown
            return f"{int(h):02d}:{mins}"
        return f"{int(h):02d}:{mins}{meridiem}"
    # already formatted, just strip
    return s if s else None


def normalize_day(val):
    if pd.isna(val):
        return None
    s = str(val).strip().lower()
    return DAY_MAP.get(s, str(val).strip() if str(val).strip() else None)


def normalize_bu(val):
    """Strip letters/symbols that are OCR noise from Business Unit codes."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    # Remove trailing letter/symbol noise like '107871-5' -> '107871.5'
    # Replace dashes used as decimal separators
    s = s.replace("-", ".").replace(" ", "")
    # Remove stray letters unless the whole thing is letters
    cleaned = re.sub(r'[A-Za-z]', '', s)
    if cleaned and cleaned not in (".", ".."):
        return cleaned.rstrip(".")
    return str(val).strip()  # fallback: keep original


def normalize_hours(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "-"):
        return None
    try:
        return float(s.replace(":", ".").replace(",", "."))
    except ValueError:
        return None


def normalize_yesno(val):
    if pd.isna(val):
        return None
    s = str(val).strip().lower()
    if s in ("yes", "y", "true", "1", "si", "sí"):
        return "Yes"
    if s in ("no", "n", "false", "0"):
        return "No"
    return None


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # ── Rename extractor columns → cleaned schema ────────────────────────────
    rename = {
        "Sched Hours": "Sched Hours",      # already correct
        "Actual Hours": "Actual Hours",
    }
    df.rename(columns=rename, inplace=True)

    # Add Recorded Hours if missing (copy from Sched Hours as default)
    if "Recorded Hours" not in df.columns:
        df["Recorded Hours"] = df.get("Sched Hours")

    # ── Ensure all output columns exist ─────────────────────────────────────
    for col in CLEAN_COLS:
        if col not in df.columns:
            df[col] = None

    # ── String strip ─────────────────────────────────────────────────────────
    str_cols = ["Source File", "Project", "Employee Name", "Title",
                "Staff Type", "Job Task", "Status"]
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda x: str(x).strip() if pd.notna(x) and str(x).strip() not in ("nan", "") else None
            )

    # ── Day normalization ─────────────────────────────────────────────────────
    df["Day"] = df["Day"].apply(normalize_day)

    # ── Business Unit ─────────────────────────────────────────────────────────
    df["Business Unit"] = df["Business Unit"].apply(normalize_bu)

    # ── Time columns ─────────────────────────────────────────────────────────
    time_cols = ["Sched Start", "Sched End", "Actual Start",
                 "Lunch Out", "Lunch In", "Actual End"]
    for col in time_cols:
        if col in df.columns:
            df[col] = df[col].apply(normalize_time)

    # ── Numeric hours ─────────────────────────────────────────────────────────
    for col in ["Sched Hours", "Actual Hours", "Recorded Hours"]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_hours)

    # ── Yes/No ────────────────────────────────────────────────────────────────
    for col in ["Absent", "Schedule Changed"]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_yesno)

    # ── EIN: strip leading zeros, keep as string ─────────────────────────────
    def clean_ein(val):
        if pd.isna(val):
            return None
        s = str(val).strip().replace(".0", "")
        return s if s and s != "nan" else None
    df["EIN"] = df["EIN"].apply(clean_ein)

    # ── Confidence: ensure float ──────────────────────────────────────────────
    df["Confidence"] = pd.to_numeric(df["Confidence"], errors="coerce")

    # ── Status: recompute from confidence ────────────────────────────────────
    def status(row):
        c = row.get("Confidence")
        if pd.isna(c):
            return "REVIEW"
        return "OK" if float(c) >= 0.7 else "Low Confidence"
    df["Status"] = df.apply(status, axis=1)

    # ── Drop rows with no employee name AND no EIN ───────────────────────────
    df = df[~(df["Employee Name"].isna() & df["EIN"].isna() & (df["Confidence"].fillna(0) < 0.4))]

    return df[CLEAN_COLS].reset_index(drop=True)


# ── Excel writer ──────────────────────────────────────────────────────────────
def write_cleaned_excel(df: pd.DataFrame, out_path: Path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned"

    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F3864")
    alt_fill = PatternFill("solid", start_color="EEF2F7")
    flag_fill = PatternFill("solid", start_color="FFF2CC")
    review_fill = PatternFill("solid", start_color="FFE0E0")

    col_widths = [
        22, 6, 12, 10, 14, 30, 14, 24, 14, 13,
        12, 10, 10, 10, 11, 11, 10, 10, 10, 11,
        8, 16, 11, 16
    ]

    # Header
    for c, col in enumerate(CLEAN_COLS, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
    ws.row_dimensions[1].height = 28

    # Data
    for r, row in enumerate(df.itertuples(index=False), start=2):
        conf = getattr(row, "Confidence", 1.0)
        status = getattr(row, "Status", "OK")
        if status == "Low Confidence":
            fill = flag_fill
        elif status == "REVIEW":
            fill = review_fill
        elif r % 2 == 0:
            fill = alt_fill
        else:
            fill = None

        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val if val is not None else "")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bdr
            if fill:
                cell.fill = fill

    # Column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"  ✅ Saved → {out_path.name}  ({len(df)} records)")


# ── File picker ───────────────────────────────────────────────────────────────
def pick_file(search_dir: Path) -> Path:
    candidates = []
    for pattern in ("*.xlsx", "*.XLSX"):
        candidates.extend(search_dir.glob(pattern))
    # deduplicate case-insensitively (Windows)
    seen = set()
    unique = []
    for f in sorted(candidates):
        if f.name.lower() not in seen:
            seen.add(f.name.lower())
            unique.append(f)

    if not unique:
        print(f"No .xlsx files found in: {search_dir}")
        sys.exit(0)

    print("\n" + "=" * 60)
    print("  TIMESHEET CLEANER — FILE SELECTION")
    print("=" * 60)
    print(f"\n  Looking in: {search_dir}\n")
    for i, f in enumerate(unique, 1):
        kb = f.stat().st_size / 1024
        print(f"  [{i}] {f.name:<45} ({kb:.0f} KB)")

    print("\nSelect file to clean (number or q to quit):")
    while True:
        raw = input("  > ").strip().lower()
        if raw == "q":
            print("Exiting.")
            sys.exit(0)
        try:
            idx = int(raw) - 1
            if 0 <= idx < len(unique):
                return unique[idx]
            print("  Out of range. Try again.")
        except ValueError:
            print("  Enter a number.")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Clean extracted timesheet Excel (no API)")
    parser.add_argument("--input",      default=None,           help="Path to input Excel file")
    parser.add_argument("--search-dir", default="data/2.output", help="Directory to search for Excel files (default: data/2.output)")
    parser.add_argument("--output-dir", default="data/3.archive", help="Directory for cleaned output (default: data/3.archive)")
    args = parser.parse_args()

    # Resolve paths relative to script location
    base = Path(__file__).parent.parent  # scripts/ -> project root

    if args.input:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"File not found: {input_path}")
            sys.exit(1)
    else:
        search_dir = base / args.search_dir
        if not search_dir.exists():
            # fallback: look in current directory
            search_dir = Path.cwd()
        input_path = pick_file(search_dir)

    output_dir = base / args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n  Loading: {input_path.name} ...")
    try:
        xl = pd.read_excel(input_path, sheet_name="All Records")
    except Exception as e:
        print(f"  ERROR reading file: {e}")
        print("  Make sure the file has an 'All Records' sheet.")
        sys.exit(1)

    print(f"  Rows loaded  : {len(xl)}")
    print(f"  Cleaning ...")

    cleaned = clean_df(xl)

    low_conf = (cleaned["Status"] == "Low Confidence").sum()
    review   = (cleaned["Status"] == "REVIEW").sum()
    print(f"  Rows after clean : {len(cleaned)}")
    print(f"  Low confidence   : {low_conf}")
    print(f"  Review flagged   : {review}")

    stem = input_path.stem
    # Avoid double _cleaned suffix
    if stem.endswith("_cleaned"):
        out_name = f"{stem}.xlsx"
    else:
        out_name = f"{stem}_cleaned.xlsx"

    out_path = output_dir / out_name
    write_cleaned_excel(cleaned, out_path)
    print(f"\n  Output dir: {output_dir}")
    print("\n✅ Done.")


if __name__ == "__main__":
    main()